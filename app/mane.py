import os
import json
import imaplib
import email
import smtplib
import threading
import time
import tempfile
from datetime import datetime, timezone
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import decode_header
import requests
from openai import OpenAI
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
import fitz
import docx
import openpyxl
import sys
from dotenv import load_dotenv
import os
load_dotenv()

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "backend"))
from database import SessionLocal, Grupo, EmailGrupo, Processo

# ============================================================
# CONFIGURAÃ‡Ã•ES
# ============================================================
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
EMAIL_HOST = "mail.realpublicidade.com.br"
EMAIL_USER = os.getenv("EMAIL_USER")
EMAIL_PASS = os.getenv("EMAIL_PASS")
EMAIL_PORT_IMAP = 993
EMAIL_PORT_SMTP = 587
DEEPSEEK_KEY = os.getenv("DEEPSEEK_KEY")

BASE = os.path.dirname(os.path.abspath(__file__))
DADOS = os.path.join(BASE, "..", "dados")
CHAT_ID_FILE = os.path.join(DADOS, "chat_id.json")
TOKEN_CALENDAR = os.path.join(DADOS, "token.json")
CONHECIMENTO_FILE = os.path.join(DADOS, "conhecimento_registro.json")

os.makedirs(DADOS, exist_ok=True)

COMANDOS_PLANILHA = ["planilhar:", "criar planilha", "organizar pasta", "planilha enel"]
COMANDOS_GRUPO = ["criar grupo empresarial:", "criar grupo:", "novo grupo empresarial:"]	

# ============================================================
# CONHECIMENTO DE REGISTRO
# ============================================================
def carregar_conhecimento():
    if os.path.exists(CONHECIMENTO_FILE):
        with open(CONHECIMENTO_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

CONHECIMENTO_REGISTRO = carregar_conhecimento()

# ============================================================
# DEEPSEEK
# ============================================================
client = OpenAI(api_key=DEEPSEEK_KEY, base_url="https://api.deepseek.com")

SYSTEM_MANE = """VocÃª Ã© o ManÃ©, assistente virtual pessoal e executivo especializado em registro empresarial na Junta Comercial.

VocÃª tem conhecimento completo dos Manuais de Registro de Sociedade AnÃ´nima e de Ltda do DREI, e atua como suporte ao processo de registro de atas e alteraÃ§Ãµes contratuais nas Juntas Comerciais brasileiras, com foco no Rio de Janeiro (JUCERJ).

SUAS ESPECIALIDADES:
1. Identificar o tipo de ato em uma ata (AGO, AGE, AGOE, RCA, ReuniÃ£o de Diretoria, Assembleia Especial, AlteraÃ§Ã£o Contratual, ARS, Escritura de DebÃªntures etc.)
2. Identificar os eventos contidos na ata (alteraÃ§Ã£o de endereÃ§o, objeto social, capital social, diretores, conselho etc.)
3. Gerar checklist completo de documentos necessÃ¡rios para registro na Junta Comercial
4. Alertar quando o processo exige CPL (Consulta PrÃ©via de Local) na Prefeitura antes da Viabilidade e DBE
5. Orientar sobre o fluxo correto de cada tipo de registro
6. Acompanhar status de processos e protocolos

REGRAS DE ANÃLISE DE ATA:
Quando receber uma ata, sempre faÃ§a:
1. IDENTIFICAÃ‡ÃƒO: Qual o tipo de ato? (AGO, AGE, AGOE, RCA etc.) â€” S/A ou Ltda?
2. EVENTOS: Quais matÃ©rias foram deliberadas? Liste todos os eventos identificados
3. CPL NECESSÃRIA? Se houver alteraÃ§Ã£o de endereÃ§o de sede ou objeto social, alertar imediatamente que Ã© necessÃ¡ria CPL na Prefeitura antes de qualquer outra providÃªncia
4. CHECKLIST: Liste todos os documentos necessÃ¡rios para o registro
5. FLUXO: Qual o fluxo correto para este processo?
6. ALERTAS: HÃ¡ alguma irregularidade ou ponto de atenÃ§Ã£o na ata?

FORMATO DE ENTREGA:
- Use bullet points e negrito para informaÃ§Ãµes crÃ­ticas
- Sempre destaque em [ATENÃ‡ÃƒO] quando houver CPL necessÃ¡ria
- Use [ALTA], [MÃ‰DIA] e [BAIXA] para prioridades
- Seja direto e executivo

REGRA ABSOLUTA: NUNCA envie emails sem aprovaÃ§Ã£o explÃ­cita.

Responda sempre em portuguÃªs brasileiro."""

# ============================================================
# UTILITÃRIOS
# ============================================================
def carregar_json(arquivo):
    if os.path.exists(arquivo):
        with open(arquivo, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def salvar_json(arquivo, dados):
    with open(arquivo, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

def get_chat_id():
    return carregar_json(CHAT_ID_FILE).get("chat_id", "")

def salvar_chat_id(chat_id):
    salvar_json(CHAT_ID_FILE, {"chat_id": str(chat_id)})

# ============================================================
# TELEGRAM
# ============================================================
def enviar_telegram(texto, chat_id=None):
    if not chat_id:
        chat_id = get_chat_id()
    if not chat_id:
        return
    try:
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
            json={"chat_id": chat_id, "text": texto[:4000], "parse_mode": "HTML"},
            timeout=10
        )
    except Exception as e:
        print(f"Erro Telegram: {e}")

def baixar_arquivo_telegram(file_id):
    try:
        resp = requests.get(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/getFile",
            params={"file_id": file_id},
            timeout=10
        )
        file_path = resp.json()["result"]["file_path"]
        url = f"https://api.telegram.org/file/bot{TELEGRAM_TOKEN}/{file_path}"
        conteudo = requests.get(url, timeout=30).content
        return conteudo, file_path.split("/")[-1]
    except Exception as e:
        print(f"Erro download arquivo: {e}")
        return None, None

# ============================================================
# EXTRAÃ‡ÃƒO DE DOCUMENTOS
# ============================================================
def extrair_pdf(conteudo_bytes):
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as f:
            f.write(conteudo_bytes)
            tmp = f.name
        doc = fitz.open(tmp)
        texto = ""
        for page in doc:
            texto += page.get_text()
        doc.close()
        os.unlink(tmp)
        return texto[:3000]
    except Exception as e:
        return f"Erro ao ler PDF: {e}"

def extrair_word(conteudo_bytes):
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".docx") as f:
            f.write(conteudo_bytes)
            tmp = f.name
        doc = docx.Document(tmp)
        texto = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
        os.unlink(tmp)
        return texto[:3000]
    except Exception as e:
        return f"Erro ao ler Word: {e}"

def extrair_excel(conteudo_bytes):
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
            f.write(conteudo_bytes)
            tmp = f.name
        wb = openpyxl.load_workbook(tmp, data_only=True)
        texto = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            texto += f"\n--- Planilha: {sheet} ---\n"
            for row in ws.iter_rows(values_only=True):
                linha = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in linha):
                    texto += " | ".join(linha) + "\n"
        os.unlink(tmp)
        return texto[:3000]
    except Exception as e:
        return f"Erro ao ler Excel: {e}"

def processar_documento(conteudo_bytes, nome_arquivo):
    nome_lower = nome_arquivo.lower()
    if nome_lower.endswith(".pdf"):
        return "PDF", extrair_pdf(conteudo_bytes)
    elif nome_lower.endswith(".docx") or nome_lower.endswith(".doc"):
        return "Word", extrair_word(conteudo_bytes)
    elif nome_lower.endswith(".xlsx") or nome_lower.endswith(".xls"):
        return "Excel", extrair_excel(conteudo_bytes)
    return None

# ============================================================
# IA
# ============================================================
historico_conversa = []

def perguntar_ai(mensagem):
    global historico_conversa
    contexto = f"\n\nCONHECIMENTO BASE DE REGISTRO:\n{json.dumps(CONHECIMENTO_REGISTRO, ensure_ascii=False)[:3000]}"
    mensagens = [{"role": "system", "content": SYSTEM_MANE + contexto}]
    mensagens += historico_conversa[-10:]
    mensagens.append({"role": "user", "content": mensagem})
    try:
        resposta = client.chat.completions.create(
            model="deepseek-chat",
            messages=mensagens,
            max_tokens=1500,
            temperature=0.7
        )
        texto = resposta.choices[0].message.content
        historico_conversa.append({"role": "user", "content": mensagem})
        historico_conversa.append({"role": "assistant", "content": texto})
        if len(historico_conversa) > 20:
            historico_conversa = historico_conversa[-20:]
        return texto
    except Exception as e:
        return f"Erro IA: {e}"

# ============================================================
# GOOGLE CALENDAR
# ============================================================
def buscar_eventos_hoje():
    try:
        creds = Credentials.from_authorized_user_file(TOKEN_CALENDAR)
        service = build('calendar', 'v3', credentials=creds)
        agora = datetime.now(timezone.utc).isoformat()
        fim_dia = datetime.now(timezone.utc).replace(hour=23, minute=59, second=59).isoformat()
        eventos = service.events().list(
            calendarId='primary',
            timeMin=agora,
            timeMax=fim_dia,
            singleEvents=True,
            orderBy='startTime'
        ).execute()
        return eventos.get('items', [])
    except Exception as e:
        print(f"Erro Calendar: {e}")
        return None

# ============================================================
# EMAIL
# ============================================================
def verificar_emails(ler_anexos=False):
    try:
        mail = imaplib.IMAP4_SSL(EMAIL_HOST, EMAIL_PORT_IMAP, timeout=15)
        mail.login(EMAIL_USER, EMAIL_PASS)
        mail.select("INBOX")
        _, mensagens = mail.search(None, "UNSEEN")
        ids = mensagens[0].split()
        if not ids:
            mail.logout()
            return []
        emails = []
        for uid in ids[-10:]:
            _, data = mail.fetch(uid, "(RFC822)")
            msg = email.message_from_bytes(data[0][1])
            assunto_raw = decode_header(msg["Subject"])[0][0]
            assunto = assunto_raw.decode() if isinstance(assunto_raw, bytes) else str(assunto_raw)
            remetente = msg.get("From", "")
            corpo = ""
            anexos = []
            conteudos_anexos = []
            _anexos_bytes = []
            if msg.is_multipart():
                for part in msg.walk():
                    if part.get_content_type() == "text/plain" and not corpo:
                        corpo = part.get_payload(decode=True).decode(errors="ignore")[:300]
                    fname = part.get_filename()
                    if fname:
                        try:
                            partes_nome = decode_header(fname)
                            fname = "".join(
                                (t.decode(enc or "utf-8", errors="ignore") if isinstance(t, bytes) else t)
                                for t, enc in partes_nome
                            )
                        except Exception:
                            pass
                        anexos.append(fname)
                        _raw = part.get_payload(decode=True)
                        if _raw:
                            _anexos_bytes.append({"nome": fname, "bytes": _raw})
                        if ler_anexos:
                            conteudo_bytes = part.get_payload(decode=True)
                            if conteudo_bytes:
                                resultado = processar_documento(conteudo_bytes, fname)
                                if resultado:
                                    tipo, texto = resultado
                                    conteudos_anexos.append({
                                        "nome": fname,
                                        "tipo": tipo,
                                        "texto": texto
                                    })
            else:
                corpo = msg.get_payload(decode=True).decode(errors="ignore")[:300]
            emails.append({
                "uid": uid.decode(),
                "assunto": assunto,
                "remetente": remetente,
                "corpo": corpo,
                "anexos": anexos,
                "conteudos_anexos": conteudos_anexos,
                "_anexos_bytes": _anexos_bytes
            })
        mail.logout()
        return emails
    except Exception as e:
        print(f"Erro email IMAP: {e}")
        return None

def enviar_email(destinatario, assunto, corpo):
    try:
        msg = MIMEMultipart()
        msg["From"] = EMAIL_USER
        msg["To"] = destinatario
        msg["Subject"] = assunto
        msg.attach(MIMEText(corpo, "plain"))
        server = smtplib.SMTP(EMAIL_HOST, EMAIL_PORT_SMTP)
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASS)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(f"Erro envio: {e}")
        return False

# ============================================================
# MONITOR AUTOMÃTICO
# ============================================================
def monitor_emails():
    print("Monitor de emails iniciado...")
    while True:
        time.sleep(3600)
        try:
            chat_id = get_chat_id()
            if not chat_id:
                continue
            emails = verificar_emails()
            if emails:
                resumo = f"ðŸ“§ <b>{len(emails)} email(s) nÃ£o lido(s):</b>\n\n"
                for e in emails:
                    resumo += f"â€¢ <b>{e['assunto']}</b>\n  De: {e['remetente']}\n"
                    if e['anexos']:
                        resumo += f"  ðŸ“Ž Anexos: {', '.join(e['anexos'])}\n"
                    resumo += "\n"
                resumo += "Quer que eu analise algum?"
                enviar_telegram(resumo, chat_id)
        except Exception as e:
            print(f"Erro monitor: {e}")

def monitor_agenda():
    print("Monitor de agenda iniciado...")
    while True:
        agora = datetime.now()
        if agora.hour == 8 and agora.minute == 0:
            try:
                chat_id = get_chat_id()
                if chat_id:
                    eventos = buscar_eventos_hoje()
                    if eventos:
                        resumo = "ðŸ“… <b>Bom dia! Sua agenda de hoje:</b>\n\n"
                        for e in eventos:
                            inicio = e['start'].get('dateTime', e['start'].get('date', ''))
                            hora = inicio[11:16] if 'T' in inicio else "Dia todo"
                            resumo += f"â€¢ <b>{hora}</b> â€” {e.get('summary', 'Sem tÃ­tulo')}\n"
                        enviar_telegram(resumo, chat_id)
                    else:
                        enviar_telegram("ðŸ“… Bom dia! Nenhum compromisso agendado para hoje.", chat_id)
            except Exception as e:
                print(f"Erro monitor agenda: {e}")
        time.sleep(60)

# ============================================================
# PROCESSADOR DE MENSAGENS
# ============================================================
email_pendente = {}

COMANDOS_AGENDA = [
    "minha agenda", "o que tenho hoje", "o que tem hoje",
    "tem algo essa semana", "ver agenda", "meu dia",
    "minha semana", "compromissos de hoje", "reunioes de hoje"
]

COMANDOS_EMAIL = [
    "verificar email", "verificar e-mail", "checar email", "checar e-mail",
    "emails novos", "tem email", "tem e-mail", "olha meu email", "olha meu e-mail",
    "meu email", "meus emails"
]

COMANDOS_ANEXOS = [
    "ler anexos", "analisar anexos", "anexos do email",
    "leia os anexos", "analisa os anexos", "documentos do email",
    "arquivos do email", "anexos novos",
    "planilhar pasta", "planilha pasta", "gerar planilha",
    "criar planilha", "planilhar:", "planilha:"
]

def extrair_email_puro(remetente):
    # remetente pode vir como 'Nome <email@x.com>' ou 'email@x.com'
    import re
    m = re.search(r'<([^>]+)>', remetente or "")
    if m:
        return m.group(1).strip().lower()
    m = re.search(r'[\w\.\-\+]+@[\w\.\-]+', remetente or "")
    return m.group(0).strip().lower() if m else ""

def grupo_do_remetente(remetente):
    email_puro = extrair_email_puro(remetente)
    if not email_puro:
        return None, None
    db = SessionLocal()
    try:
        assoc = db.query(EmailGrupo).filter(EmailGrupo.email == email_puro).first()
        if assoc:
            return assoc.grupo_id, email_puro
        return None, email_puro
    finally:
        db.close()
UPLOADS_BACKEND = os.path.join(BASE, "..", "backend", "uploads")

ASSUNTOS_PROCESSO = ["processo para registro", "processos para registro"]

def analisar_ata_para_processo(texto):
    prompt = f"""Analise o preambulo deste documento societario e extraia em JSON.

DOCUMENTO:
{texto[:3000]}

Regras:
- "empresa": razao social completa, sem CNPJ.
- "cnpj": formato XX.XXX.XXX/XXXX-XX. Se nao achar, "".
- "nire": numero do NIRE se houver, senao "".
- "tipo_sociedade": "SA" ou "LTDA".
- "tipo_ato": ex AGO, AGE, AGOE, RCA, ALTERACAO_CONTRATUAL.
- "identificador_ato": descricao por extenso com a data, ex "ATA DE REUNIAO DE SOCIOS REALIZADA EM 29 DE ABRIL DE 2026".
- "data_ata": "DD/MM/AAAA".
- "uf": sigla de 2 letras do estado da SEDE da empresa (ex: RJ, SP, BA). Se so houver o nome por extenso, devolva a sigla. Se nao achar, "".

Retorne APENAS este JSON:
{{"empresa":"","cnpj":"","nire":"","tipo_sociedade":"","tipo_ato":"","identificador_ato":"","data_ata":"","uf":""}}"""
    resp = client.chat.completions.create(
        model="deepseek-chat",
        messages=[{"role": "user", "content": prompt}],
        max_tokens=400,
        temperature=0.1
    )
    txt = resp.choices[0].message.content.replace("```json","").replace("```","").strip()
    return json.loads(txt)

def criar_processo_de_anexo(conteudo_bytes, nome_arquivo, grupo_id):
    import uuid as _uuid
    from datetime import datetime as _dt
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as f:
        f.write(conteudo_bytes)
        tmp = f.name
    doc = fitz.open(tmp)
    texto = ""
    for page in doc:
        texto += page.get_text()
    doc.close()
    os.unlink(tmp)

    dados = analisar_ata_para_processo(texto)

    processo_id = f"MN-{_dt.now().strftime('%Y%m%d%H%M%S')}-{str(_uuid.uuid4())[:4].upper()}"
    os.makedirs(UPLOADS_BACKEND, exist_ok=True)
    nome_salvo = f"{processo_id}_ata.pdf"
    with open(os.path.join(UPLOADS_BACKEND, nome_salvo), "wb") as f:
        f.write(conteudo_bytes)

    db = SessionLocal()
    try:
        p = Processo(
            id=processo_id,
            empresa=dados.get("empresa","") or "",
            cnpj=dados.get("cnpj","") or "",
            nire=dados.get("nire","") or "",
            tipo_sociedade=dados.get("tipo_sociedade","") or "",
            tipo_ato=dados.get("tipo_ato","") or "",
            identificador_ato=dados.get("identificador_ato","") or "",
            data_ata=dados.get("data_ata","") or "",
            uf=(dados.get("uf","") or "").strip().upper()[:2],
            status="recebido",
            arquivo_ata=nome_salvo,
            grupo_id=grupo_id
        )
        db.add(p)
        db.commit()
        return p.empresa or nome_arquivo
    finally:
        db.close()

def processar_email_para_processos(email_obj):
    assunto = (email_obj.get("assunto","") or "").strip().lower()
    if assunto not in ASSUNTOS_PROCESSO:
        return None
    grupo_id, email_puro = grupo_do_remetente(email_obj.get("remetente",""))
    if not grupo_id:
        return f"Email de {email_puro or 'desconhecido'} com assunto de processo, mas remetente nao cadastrado."
    criados = []
    for anexo in email_obj.get("_anexos_bytes", []):
        nome = anexo["nome"]
        if not nome.lower().endswith(".pdf"):
            continue
        try:
            empresa = criar_processo_de_anexo(anexo["bytes"], nome, grupo_id)
            criados.append(empresa)
        except Exception as e:
            criados.append(f"ERRO em {nome}: {e}")
    if not criados:
        return f"Email de {email_puro}: nenhum PDF processado."
    return f"{len(criados)} processo(s) criado(s) para o grupo do remetente {email_puro}:\n- " + "\n- ".join(criados)
def processar_mensagem(texto, chat_id):
    global email_pendente
    salvar_chat_id(chat_id)
    texto_lower = texto.lower().strip()

    if email_pendente:
        if texto_lower in ["sim", "confirmar", "enviar", "ok", "aprovado", "pode enviar"]:
            sucesso = enviar_email(
                email_pendente["destinatario"],
                email_pendente["assunto"],
                email_pendente["corpo"]
            )
            email_pendente = {}
            return "âœ… Email enviado!" if sucesso else "âŒ Erro ao enviar. Tente novamente."
        elif texto_lower in ["nÃ£o", "nao", "cancelar", "cancel", "nÃ£o enviar"]:
            email_pendente = {}
            return "âŒ Envio cancelado."

    if any(p in texto_lower for p in COMANDOS_AGENDA):
        print("Verificando agenda...")
        eventos = buscar_eventos_hoje()
        if eventos is None:
            return "âŒ Erro ao conectar no Google Calendar."
        if not eventos:
            return "ðŸ“… Nenhum compromisso para hoje."
        resumo = "ðŸ“… <b>Sua agenda de hoje:</b>\n\n"
        for e in eventos:
            inicio = e['start'].get('dateTime', e['start'].get('date', ''))
            hora = inicio[11:16] if 'T' in inicio else "Dia todo"
            resumo += f"â€¢ <b>{hora}</b> â€” {e.get('summary', 'Sem tÃ­tulo')}\n"
        return resumo
        
    if texto_lower.startswith("cadastrar cliente:"):
        resto = texto.split(":", 1)[-1].strip()
        partes = resto.split()
        email_cliente = ""
        for p in partes:
            if "@" in p:
                email_cliente = p
                break
        if not email_cliente:
            return "Informe o email. Ex: Cadastrar Cliente: NOME DO GRUPO email@cliente.com"
        nome_grupo = resto.replace(email_cliente, "").strip().upper()
        if not nome_grupo:
            return "Informe o nome do grupo. Ex: Cadastrar Cliente: NOME DO GRUPO email@cliente.com"
        db = SessionLocal()
        try:
            import uuid
            grupo = db.query(Grupo).filter(Grupo.nome == nome_grupo).first()
            if not grupo:
                base = "".join(c for c in nome_grupo if c.isalnum())[:10].upper()
                codigo = f"{base}-{uuid.uuid4().hex[:4].upper()}"
                grupo = Grupo(id=str(uuid.uuid4()), nome=nome_grupo, codigo=codigo)
                db.add(grupo)
                db.commit()
            assoc = db.query(EmailGrupo).filter(EmailGrupo.email == email_cliente, EmailGrupo.grupo_id == grupo.id).first()
            if not assoc:
                db.add(EmailGrupo(id=str(uuid.uuid4()), email=email_cliente, grupo_id=grupo.id))
                db.commit()
            link = f"http://localhost:3000/cliente?grupo={grupo.codigo}"
            corpo = (
                f"Ola,\n\nVoce foi convidado a acessar o sistema do grupo {grupo.nome}.\n\n"
                f"Para criar seu login e senha, acesse o link abaixo:\n{link}\n\n"
                f"Apos o cadastro, voce podera acompanhar e baixar seus documentos.\n\nMane"
            )
            enviou = enviar_email(email_cliente, f"Acesso ao sistema - {grupo.nome}", corpo)
            if enviou:
                return f"Cliente cadastrado!\n\nGrupo: <b>{grupo.nome}</b>\nCodigo: <code>{grupo.codigo}</code>\nEmail enviado para: {email_cliente}"
            else:
                return f"Grupo {grupo.nome} pronto (codigo {grupo.codigo}), mas falhou o envio do email para {email_cliente}. Verifique as credenciais de email."
        except Exception as e:
            db.rollback()
            return f"Erro ao cadastrar cliente: {e}"
        finally:
            db.close()
    if any(p in texto_lower for p in COMANDOS_GRUPO):
        nome_grupo = texto.split(":", 1)[-1].strip().upper()
        if not nome_grupo:
            return "âŒ Informe o nome do grupo. Ex: Criar grupo empresarial: HIG CAPITAL"
        db = SessionLocal()
        try:
            existente = db.query(Grupo).filter(Grupo.nome == nome_grupo).first()
            if existente:
                return f"âš ï¸ JÃ¡ existe um grupo chamado <b>{nome_grupo}</b>.\nCÃ³digo: <code>{existente.codigo}</code>"
            import uuid
            base = "".join(c for c in nome_grupo if c.isalnum())[:10].upper()
            sufixo = uuid.uuid4().hex[:4].upper()
            codigo = f"{base}-{sufixo}"
            novo = Grupo(id=str(uuid.uuid4()), nome=nome_grupo, codigo=codigo)
            db.add(novo)
            db.commit()
            return f"âœ… Grupo criado!\n\nðŸ¢ <b>{nome_grupo}</b>\nðŸ”‘ CÃ³digo: <code>{codigo}</code>\n\nEsse cÃ³digo vai no link de cadastro do cliente."
        except Exception as e:
            db.rollback()
            return f"âŒ Erro ao criar grupo: {e}"
        finally:
            db.close()
# Planilhar pasta
    if any(p in texto_lower for p in COMANDOS_PLANILHA):
        pasta = texto.split(":")[-1].strip().strip('"').strip("'")
        if not os.path.exists(pasta):
            return f"âŒ Pasta nÃ£o encontrada: {pasta}"
        enviar_telegram(f"ðŸ“Š Iniciando planilhamento da pasta...\n{pasta}", chat_id)
        def planilhar_em_background(pasta, chat_id):
            try:
                import fitz
                from openai import OpenAI
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                deepseek = OpenAI(api_key=DEEPSEEK_KEY, base_url="https://api.deepseek.com")
                pdfs = sorted([f for f in os.listdir(pasta) if f.lower().endswith(".pdf")])
                if not pdfs:
                    enviar_telegram("âŒ Nenhum PDF encontrado na pasta.", chat_id)
                    return
                enviar_telegram(f"ðŸ“„ Encontrei {len(pdfs)} PDF(s). Analisando...", chat_id)
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Processos"
                cabecalho = ["EMPRESA", "ATO", "DATA", "STATUS"]
                for col, titulo in enumerate(cabecalho, 1):
                    cell = ws.cell(row=1, column=col, value=titulo)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(fill_type="solid", fgColor="1E3A5F")
                    cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions["A"].width = 55
                ws.column_dimensions["B"].width = 35
                ws.column_dimensions["C"].width = 20
                ws.column_dimensions["D"].width = 30
                for i, pdf in enumerate(pdfs, 2):
                    caminho = os.path.join(pasta, pdf)
                    try:
                        doc = fitz.open(caminho)
                        texto_pdf = ""
                        for page in doc:
                            texto_pdf += page.get_text()
                        doc.close()
                        texto_pdf = texto_pdf[:2000]
                        prompt = f"""Analise este documento e extraia APENAS estas informaÃ§Ãµes em JSON:
{texto_pdf}
Retorne APENAS este JSON:
{{"empresa": "nome completo da empresa", "ato": "tipo de ato ex: AGOE, 1Âª AlteraÃ§Ã£o Contratual", "data": "DD/MM/AAAA"}}"""
                        resp = deepseek.chat.completions.create(
                            model="deepseek-chat",
                            messages=[{"role": "user", "content": prompt}],
                            max_tokens=200,
                            temperature=0.1
                        )
                        import json
                        dados = json.loads(resp.choices[0].message.content.replace("```json","").replace("```","").strip())
                        ws.cell(row=i, column=1, value=dados.get("empresa","").upper())
                        ws.cell(row=i, column=2, value=dados.get("ato","").upper())
                        ws.cell(row=i, column=3, value=dados.get("data",""))
                        ws.cell(row=i, column=4, value="TRAMITANDO")
                    except Exception as e:
                        ws.cell(row=i, column=1, value=pdf)
                        ws.cell(row=i, column=2, value="ERRO AO LER")
                        ws.cell(row=i, column=4, value="TRAMITANDO")
                nome_planilha = f"processos_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
                saida = os.path.join(pasta, nome_planilha)
                wb.save(saida)
                enviar_telegram(f"âœ… Planilha criada com {len(pdfs)} empresas!\nðŸ“ Salva em: {saida}", chat_id)
            except Exception as e:
                enviar_telegram(f"âŒ Erro: {e}", chat_id)
        threading.Thread(target=planilhar_em_background, args=(pasta, chat_id), daemon=True).start()
        return "â³ Processando..."

    if any(p in texto_lower for p in COMANDOS_ANEXOS):
        enviar_telegram("ðŸ“Ž Verificando seus emails com anexos... Aguarde.", chat_id)
        def processar_em_background(chat_id):
            emails = verificar_emails(ler_anexos=True)
            if emails is None:
                enviar_telegram("âŒ Erro ao conectar no servidor de email.", chat_id)
                return
            emails_com_anexos = [e for e in emails if e['conteudos_anexos']]
            if not emails_com_anexos:
                enviar_telegram("ðŸ“­ Nenhum email com anexos legÃ­veis no momento.", chat_id)
                return
            resumo = f"ðŸ“Ž <b>Encontrei {len(emails_com_anexos)} email(s) com anexos:</b>\n\n"
            for e in emails_com_anexos:
                resumo += f"â€¢ <b>{e['assunto']}</b> â€” {e['remetente']}\n"
                for anexo in e['conteudos_anexos']:
                    prompt = f"Email: {e['assunto']}\nAnexo {anexo['tipo']}: {anexo['nome']}\nConteÃºdo:\n{anexo['texto']}\n\nFaÃ§a um resumo executivo em 3 bullet points."
                    analise = perguntar_ai(prompt)
                    resumo += f"  ðŸ“„ <b>{anexo['nome']}</b>\n{analise}\n\n"
            enviar_telegram(resumo, chat_id)
        threading.Thread(target=processar_em_background, args=(chat_id,), daemon=True).start()
        return "â³ Processando..."

    if any(p in texto_lower for p in COMANDOS_EMAIL):
        print("Verificando emails...")
        emails = verificar_emails()
        if emails is None:
            return "âŒ Erro ao conectar no servidor de email."
        if not emails:
            return "ðŸ“­ Nenhum email nÃ£o lido no momento."
        resumo = f"ðŸ“§ <b>{len(emails)} email(s) nÃ£o lido(s):</b>\n\n"
        for e in emails:
            resumo += f"â€¢ <b>{e['assunto']}</b>\n  De: {e['remetente']}\n"
            if e['anexos']:
                resumo += f"  ðŸ“Ž {', '.join(e['anexos'])}\n"
            resumo += f"  {e['corpo'][:100]}...\n\n"
        return resumo

    return perguntar_ai(texto)

def processar_documento_telegram(update, chat_id):
    msg = update.get("message", {})
    documento = msg.get("document", {})
    file_id = documento.get("file_id")
    nome = documento.get("file_name", "arquivo")

    if not file_id:
        return

    enviar_telegram(f"ðŸ“„ Recebi <b>{nome}</b>. Analisando...", chat_id)

    conteudo, _ = baixar_arquivo_telegram(file_id)
    if not conteudo:
        enviar_telegram("âŒ Erro ao baixar o arquivo.", chat_id)
        return

    resultado = processar_documento(conteudo, nome)
    if not resultado:
        enviar_telegram("âŒ Formato nÃ£o suportado. Envie PDF, Word ou Excel.", chat_id)
        return

    tipo, texto = resultado
    prompt = f"""Recebi o seguinte documento {tipo} chamado '{nome}'.

ConteÃºdo:
{texto}

Por favor analise este documento seguindo o protocolo:
1. IDENTIFICAÃ‡ÃƒO: Que tipo de ato Ã© este? (AGO, AGE, AGOE, RCA, AlteraÃ§Ã£o Contratual, ARS etc.) â€” S/A ou Ltda?
2. EVENTOS: Quais matÃ©rias foram deliberadas?
3. CPL NECESSÃRIA? HÃ¡ alteraÃ§Ã£o de endereÃ§o ou objeto social?
4. CHECKLIST: Quais documentos sÃ£o necessÃ¡rios para o registro?
5. FLUXO: Qual o fluxo correto?
6. ALERTAS: HÃ¡ algum ponto de atenÃ§Ã£o?"""

    analise = perguntar_ai(prompt)
    enviar_telegram(f"ðŸ“„ <b>AnÃ¡lise: {nome}</b>\n\n{analise}", chat_id)

# ============================================================
# POLLING
# ============================================================
def monitor_processos():
    print("Monitor de processos (email -> sistema) iniciado...")
    INTERVALO = 30  # 30 segundos
    while True:
        time.sleep(INTERVALO)
        try:
            emails = verificar_emails(ler_anexos=False)
            if not emails:
                continue
            chat_id = get_chat_id()
            for e in emails:
                resultado = processar_email_para_processos(e)
                if resultado and chat_id:
                    enviar_telegram(resultado, chat_id)
        except Exception as ex:
            print(f"Erro monitor_processos: {ex}")
def iniciar_polling():
    print("ManÃ© iniciando...")
    ultimo_update = 0

    while True:
        try:
            resp = requests.get(
                f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/getUpdates",
                params={"timeout": 20, "offset": ultimo_update + 1},
                timeout=25
            )
            data = resp.json()

            for update in data.get("result", []):
                ultimo_update = update["update_id"]
                msg = update.get("message", {})
                chat_id = msg.get("chat", {}).get("id", "")

                if not chat_id:
                    continue

                salvar_chat_id(chat_id)

                if msg.get("document"):
                    threading.Thread(
                        target=processar_documento_telegram,
                        args=(update, chat_id),
                        daemon=True
                    ).start()
                    continue

                texto = msg.get("text", "")
                if texto:
                    print(f"Mensagem: {texto}")
                    resposta = processar_mensagem(texto, chat_id)
                    enviar_telegram(resposta, chat_id)

        except requests.exceptions.Timeout:
            pass
        except Exception as e:
            print(f"Erro: {e}")
            time.sleep(5)

# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    threading.Thread(target=monitor_emails, daemon=True).start()
    threading.Thread(target=monitor_agenda, daemon=True).start()
    threading.Thread(target=monitor_processos, daemon=True).start()
    iniciar_polling()
