from fastapi import FastAPI, Depends, UploadFile, File, Form, HTTPException, Header, BackgroundTasks, Request, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session, defer
from sqlalchemy import func
from database import get_db, Processo, Grupo, Usuario, EmailGrupo, criar_banco, AuditLog, Codigo2FA, Anexo, RegraAprendizado, MensagemProcesso, TelegramVinculo, Fluxo, Evento, AssistenteConversa
from cnpj_utils import normalizar_cnpj, validar_cnpj, formatar_cnpj
from datetime import datetime, timedelta, date
from openai import OpenAI
import json, os, uuid, shutil, bcrypt, secrets
import asyncio

from dotenv import load_dotenv
import os
load_dotenv()

# --- Rate limiter simples em memoria (anti-forca-bruta no login) ---
import time as _time
_login_tentativas = {}
_LOGIN_MAX = 5          # tentativas
_LOGIN_JANELA = 300     # segundos (5 min)
_LOGIN_BLOQUEIO = 900   # segundos (15 min de bloqueio)
def _checar_rate_login(ip):
    agora = _time.time()
    reg = _login_tentativas.get(ip)
    if reg and reg.get("bloqueado_ate", 0) > agora:
        return False
    if not reg or (agora - reg.get("inicio", 0)) > _LOGIN_JANELA:
        _login_tentativas[ip] = {"inicio": agora, "falhas": 0, "bloqueado_ate": 0}
    return True
def _registrar_falha_login(ip):
    agora = _time.time()
    reg = _login_tentativas.get(ip) or {"inicio": agora, "falhas": 0, "bloqueado_ate": 0}
    reg["falhas"] = reg.get("falhas", 0) + 1
    if reg["falhas"] >= _LOGIN_MAX:
        reg["bloqueado_ate"] = agora + _LOGIN_BLOQUEIO
    _login_tentativas[ip] = reg
def _limpar_falhas_login(ip):
    if ip in _login_tentativas:
        del _login_tentativas[ip]

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

EMAIL_USER = os.getenv("EMAIL_USER")
EMAIL_FROM = os.getenv("EMAIL_FROM") or EMAIL_USER
EMAIL_PASS = os.getenv("EMAIL_PASS")
EMAIL_HOST = os.getenv("EMAIL_HOST", "smtp.gmail.com")
EMAIL_PORT_SMTP = int(os.getenv("EMAIL_PORT_SMTP", "587"))
BASE_URL_SISTEMA = os.getenv("BASE_URL_SISTEMA", "https://atos.net.br")

def enviar_email(destinatario, assunto, corpo, corpo_html=None):
    try:
        msg = MIMEMultipart("alternative")
        msg["From"] = "Atos - Gestao Societaria <%s>" % EMAIL_FROM
        msg["To"] = destinatario
        msg["Subject"] = assunto
        msg.attach(MIMEText(corpo, "plain"))
        if not corpo_html:
            try:
                corpo_html = envolver_html(corpo)
            except Exception:
                corpo_html = None
        if corpo_html:
            msg.attach(MIMEText(corpo_html, "html"))
        server = smtplib.SMTP(EMAIL_HOST, EMAIL_PORT_SMTP)
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASS)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(f"Erro ao enviar email para {destinatario}: {e}")
        return False

def enviar_email_anexo(destinatario, assunto, corpo, caminho_anexo=None, nome_anexo=None, corpo_html=None):
    try:
        from email.mime.application import MIMEApplication
        msg = MIMEMultipart("mixed")
        msg["From"] = "Atos - Gestao Societaria <%s>" % EMAIL_FROM
        msg["To"] = destinatario
        msg["Subject"] = assunto
        alt = MIMEMultipart("alternative")
        alt.attach(MIMEText(corpo, "plain"))
        try:
            alt.attach(MIMEText(corpo_html or envolver_html(corpo), "html"))
        except Exception as _e:
            print("aviso html anexo:", _e)
        msg.attach(alt)
        if caminho_anexo and os.path.exists(caminho_anexo):
            with open(caminho_anexo, "rb") as fa:
                part = MIMEApplication(fa.read(), Name=(nome_anexo or os.path.basename(caminho_anexo)))
            part["Content-Disposition"] = 'attachment; filename="%s"' % (nome_anexo or os.path.basename(caminho_anexo))
            msg.attach(part)
        server = smtplib.SMTP(EMAIL_HOST, EMAIL_PORT_SMTP)
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASS)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(f"Erro ao enviar email (anexo) para {destinatario}: {e}")
        return False

def validar_token(x_token, db):
    """Busca o usuario pelo token e verifica se nao expirou (30 dias)."""
    if not x_token:
        return None
    u = db.query(Usuario).filter(Usuario.token == x_token).first()
    if not u:
        return None
    tc = getattr(u, "token_criado_em", None)
    if tc is not None:
        from datetime import timedelta
        if datetime.now() - tc > timedelta(days=30):
            return None  # token expirado
    return u

def obter_ip(request):
    """Retorna o IP real do cliente, lendo cabecalhos do proxy nginx."""
    if not request:
        return None
    xr = request.headers.get("x-real-ip")
    if xr:
        return xr
    xff = request.headers.get("x-forwarded-for")
    if xff:
        return xff.split(",")[0].strip()
    return request.client.host if request.client else None

def email_do_usuario(db, usuario):
    """Retorna o e-mail do usuario (proprio se houver, senao o primeiro e-mail do grupo)."""
    try:
        e = getattr(usuario, "email", None)
        if e:
            return e
        eg = db.query(EmailGrupo).filter(EmailGrupo.grupo_id == usuario.grupo_id).first()
        return eg.email if eg else None
    except Exception as ex:
        print("Erro email_do_usuario:", ex)
        return None


def registrar_auditoria(db, usuario, acao, processo_id=None, detalhe=None, ip=None):
    """Registra uma acao na trilha de auditoria. Nunca quebra a operacao principal."""
    try:
        import uuid as _uuid
        log = AuditLog(
            id=str(_uuid.uuid4()),
            usuario_login=getattr(usuario, "login", None) if usuario else None,
            usuario_nome=nome_usuario(usuario) if usuario else None,
            usuario_papel=(getattr(usuario, "papel", None) or ("admin" if getattr(usuario, "is_admin", False) else None)) if usuario else None,
            usuario_id=getattr(usuario, "id", None) if usuario else None,
            grupo_id=getattr(usuario, "grupo_id", None) if usuario else None,
            is_admin=bool(getattr(usuario, "is_admin", False)) if usuario else False,
            acao=acao,
            processo_id=processo_id,
            detalhe=detalhe,
            ip=ip,
        )
        db.add(log)
        db.commit()
    except Exception as e:
        print("Falha ao registrar auditoria:", e)

def emails_do_grupo(db, grupo_id):
    if not grupo_id:
        return []
    regs = db.query(EmailGrupo).filter(EmailGrupo.grupo_id == grupo_id).all()
    return [r.email for r in regs if r.email]


def _tem_acesso_admin(usuario):
    """Admin OU operador: mesma tela administrativa, mesma visibilidade de
    processos de todos os grupos, mesmas acoes operacionais. Cliente nao passa."""
    return bool(usuario and (usuario.is_admin or getattr(usuario, "papel", None) == "operador"))


def requer_acesso_admin(usuario):
    """Levanta 403 se o usuario nao for admin nem operador."""
    if not _tem_acesso_admin(usuario):
        raise HTTPException(status_code=403, detail="Acesso restrito a administrador ou operador")


def requer_admin_completo(usuario):
    """Levanta 403 se o usuario nao for especificamente o papel 'admin' (superadmin).
    Operador NAO passa aqui - usar nos endpoints de configuracao do sistema,
    gerenciamento de usuarios/integracoes e acoes destrutivas (excluir processo)."""
    if not usuario or not usuario.is_admin:
        raise HTTPException(status_code=403, detail="Acesso restrito ao administrador")


def nome_usuario(usuario):
    """Nome de exibicao pra atribuicao na timeline. None (sem usuario) = automacao."""
    if not usuario:
        return "Sistema (automação)"
    return getattr(usuario, "nome", None) or usuario.login

def _regex_protocolo(texto):
    """Reconhece o numero de protocolo em texto livre (ja extraido via texto
    direto ou OCR). Cobre JUCESP (0.000.000/00-0) e JUCERJA (AAAA/NNNNNNN-N,
    com o miolo variando de 7 a 9 digitos conforme o volume de protocolos do
    ano) - com ou sem separadores, ja que o print de pagina web as vezes
    quebra a barra/hifen em espacos ou quebras de linha."""
    import re
    up = texto.upper()
    if "JUCESP PROTOCOLO" in up:
        idx = up.find("JUCESP PROTOCOLO")
        m = re.search(r"\d\.\d{3}\.\d{3}/\d{2}-\d", texto[idx: idx + 120])
        if m:
            return m.group(0)
    m = re.search(r"\d\.\d{3}\.\d{3}/\d{2}-\d", texto)
    if m:
        return m.group(0)
    if "PROTOCOLO" in up:
        idx = up.rfind("PROTOCOLO")
        janela = texto[idx: idx + 80]
        mm = re.search(r"(20\d{2})\s*/\s*([\d\s]{7,11}?)\s*-\s*(\d)", janela)
        if mm:
            meio = re.sub(r"\D", "", mm.group(2))
            if 7 <= len(meio) <= 9:
                return mm.group(1) + "/" + meio + "-" + mm.group(3)
        # variacao sem separador (barra/hifen perdidos na extracao): bloco
        # continuo de digitos comecando com 20xx (ano) - ano(4) + miolo(7-9) + dv(1)
        janela_digitos = re.sub(r"\s", "", janela)
        mm2 = re.search(r"(20\d{2})(\d{7,9})(\d)(?!\d)", janela_digitos)
        if mm2:
            return mm2.group(1) + "/" + mm2.group(2) + "-" + mm2.group(3)
    m = re.search(r"20\d{2}/\d{7,9}-\d", texto)
    if m:
        return m.group(0)
    # Ultimo fallback: o numero pode ter sido quebrado por espaco/quebra de
    # linha em qualquer lugar do texto, nao so perto da palavra "protocolo"
    # (ex: layout de print de pagina web que virou a linha no meio do numero).
    texto_sem_espaco = re.sub(r"\s", "", texto)
    m2 = re.search(r"20\d{2}/\d{7,9}-\d", texto_sem_espaco)
    if m2:
        return m2.group(0)
    return None

def _gemini_protocolo(caminho_pdf):
    import base64, json, urllib.request
    if not GEMINI_KEY:
        return None
    try:
        with open(caminho_pdf, "rb") as f:
            pdf_b64 = base64.b64encode(f.read()).decode()
        prompt = ("Verifique se este documento e um COMPROVANTE de protocolo de Junta Comercial "
                  "(JUCESP ou JUCERJA) - ou seja, a tela/recibo que confirma que um pedido acabou "
                  "de ser protocolado, geralmente com o titulo 'PROTOCOLO GERADO COM SUCESSO' ou "
                  "similar. NAO e uma ata, contrato, distrato ou qualquer outro documento societario "
                  "que apenas MENCIONE um numero de registro antigo ou de outro processo - nesses "
                  "casos, mesmo que existam numeros parecidos com protocolo no texto, eles NAO devem "
                  "ser extraidos. "
                  "Se for de fato um comprovante de protocolo, responda APENAS com o numero do "
                  "protocolo, sem mais nada. JUCESP tem formato 0.000.000/00-0. JUCERJA tem formato "
                  "2026/00000000-0 (miolo pode variar de 7 a 9 digitos). "
                  "Se o documento NAO for um comprovante de protocolo, ou se voce nao tiver certeza "
                  "de que o numero encontrado e realmente o protocolo (e nao um NIRE, CNPJ ou numero "
                  "de registro de outro ato), responda exatamente: NENHUM")
        body = {
            "contents": [{"parts": [
                {"text": prompt},
                {"inline_data": {"mime_type": "application/pdf", "data": pdf_b64}}
            ]}]
        }
        url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-flash-latest:generateContent?key=" + GEMINI_KEY
        req = urllib.request.Request(url, data=json.dumps(body).encode(), headers={"Content-Type": "application/json"})
        resp = urllib.request.urlopen(req, timeout=40)
        data = json.loads(resp.read().decode())
        txt = data["candidates"][0]["content"]["parts"][0]["text"].strip()
        if "NENHUM" in txt.upper():
            return None
        num = _regex_protocolo(txt)
        if num:
            return num
        # Resposta livre da IA que nao bate em nenhum padrao conhecido: so
        # aceita se parecer mesmo um numero de protocolo (token curto,
        # majoritariamente digitos, sem espacos) - evita aceitar como
        # protocolo uma frase/explicacao ou um numero de outro contexto
        # (NIRE, CNPJ, registro antigo mencionado no corpo de uma ata) que a
        # IA tenha confundido.
        candidato = txt.strip()
        if candidato and len(candidato) <= 30 and " " not in candidato and sum(c.isdigit() for c in candidato) >= 6:
            return candidato
        return None
    except Exception as e:
        print("Gemini protocolo falhou:", e)
        return None

def _extrair_protocolo_barcode(caminho_pdf):
    """Tenta ler o codigo de barras do protocolo direto da imagem do PDF.
    Muito mais confiavel que OCR/IA: leitura deterministica, sem risco de
    confundir digitos visualmente (ex: 6 lido como 0)."""
    import subprocess, tempfile, os, glob, re as _re
    try:
        from pyzbar.pyzbar import decode as _zbar_decode
        from PIL import Image as _PILImage
    except Exception:
        return None
    try:
        d = tempfile.mkdtemp()
        subprocess.run(["pdftoppm", "-r", "200", "-jpeg", caminho_pdf, os.path.join(d, "pg")], check=True, timeout=60)
        for img_path in sorted(glob.glob(os.path.join(d, "*.jpg"))):
            img = _PILImage.open(img_path)
            for r in _zbar_decode(img):
                digitos = _re.sub(r"\D", "", r.data.decode("utf-8", errors="ignore"))
                if len(digitos) == 10:
                    return digitos[0] + "." + digitos[1:4] + "." + digitos[4:7] + "/" + digitos[7:9] + "-" + digitos[9]
                if len(digitos) == 13:
                    return digitos[:4] + "/" + digitos[4:12] + "-" + digitos[12]
        return None
    except Exception as e:
        print("Erro ao ler codigo de barras:", str(e)[:150])
        return None


def extrair_protocolo_ocr(caminho_pdf):
    """Extrai o numero de protocolo de um comprovante (JUCESP ou JUCERJA),
    usando a mesma cadeia de tentativas de extrair_texto_pdf_em_camadas
    (texto direto via pdfplumber/fitz -> OCR via Gemini vision/pytesseract) -
    necessario porque comprovantes da JUCERJA costumam ser prints de pagina
    web salvos em PDF, sem nenhum texto selecionavel. Se nada for
    reconhecido automaticamente, retorna None e o operador preenche
    manualmente (o campo aceita texto livre, sem nenhuma restricao de
    formato)."""
    # 0) Codigo de barras: deterministico, mais confiavel que qualquer OCR/IA
    num = _extrair_protocolo_barcode(caminho_pdf)
    if num:
        print("protocolo via codigo de barras:", num)
        return num
    # 1) texto direto + 2) OCR (mesma cadeia usada pra leitura das atas)
    texto, _leitura_parcial = extrair_texto_pdf_em_camadas(caminho_pdf)
    if texto:
        num = _regex_protocolo(texto)
        if num:
            print("protocolo via texto/OCR em camadas:", num)
            return num
    # 3) Gemini com prompt dedicado (pede so o numero do protocolo) - mais
    # preciso que aplicar regex sobre o texto generico quando o layout do
    # comprovante confunde a extracao acima
    num = _gemini_protocolo(caminho_pdf)
    if num:
        print("protocolo via Gemini (prompt dedicado):", num)
    return num

def corpo_status_cliente(p, status_label, frase_final):
    ato = p.identificador_ato or p.tipo_ato or ""
    linhas = []
    linhas.append("Empresa: " + (p.empresa or ""))
    linhas.append("Ato: " + ato)
    linhas.append("Status: " + status_label)
    if p.numero_protocolo:
        linhas.append("")
        linhas.append("Protocolo: " + p.numero_protocolo)
    if frase_final:
        linhas.append("")
        linhas.append(frase_final)
    return "\n".join(linhas)





def _empresa_linha(p):
    ato = p.identificador_ato or p.tipo_ato or ""
    return (p.empresa or "") + (" · " + ato if ato else "")


def _pill_status(status_key):
    cores = {
        "aberto": ("rgba(255,159,10,0.15)", "#ff9f0a"),
        "tramitacao": ("rgba(0,212,255,0.15)", "#00d4ff"),
        "exigencia": ("rgba(255,77,77,0.15)", "#ff4d4d"),
        "deferido": ("rgba(77,148,255,0.15)", "#4d94ff"),
        "finalizado": ("rgba(0,255,170,0.15)", "#00ffaa"),
    }
    return cores.get(status_key, ("rgba(77,148,255,0.15)", "#4d94ff"))


def _email_status_html(status_key, status_label, titulo, empresa_linha, protocolo=None, nota_tipo=None, nota_texto=None, botao=None):
    """Template escuro compartilhado por todos os avisos de status ao cliente -
    replica a identidade visual aprovada em emails_notificacao_status_v2.html,
    com cores solidas e sem blur/backdrop-filter (nao renderizam em Gmail/Outlook).
    nota_tipo: None | "recebido" | "anexo" | "aguardando"."""
    bg, cor = _pill_status(status_key)
    info_html = ""
    if protocolo:
        info_html += '<div style="font-size:10.5px;color:#71717a;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:4px;">Protocolo</div>'
        info_html += '<div style="font-size:14px;color:#fff;font-weight:600;">' + protocolo + '</div>'
    elif nota_tipo == "recebido":
        info_html += '<div style="font-size:10.5px;color:#71717a;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:4px;">Recebido em</div>'
        info_html += '<div style="font-size:14px;color:#fff;font-weight:600;">' + (nota_texto or "") + '</div>'
    if nota_tipo == "anexo":
        info_html += '<div style="font-size:12px;color:#8ec2ff;margin-top:10px;">📎 ' + (nota_texto or "Documento em Anexo") + '</div>'
    elif nota_tipo == "aguardando":
        info_html += '<div style="font-size:12.5px;color:#8a90b8;margin-top:14px;line-height:1.5;">' + (nota_texto or "") + '</div>'
    botao_html = ""
    if botao:
        botao_html = ('<p style="text-align:center;margin-top:20px;"><a href="' + botao["href"] +
                       '" style="display:inline-block;background:' + cor + ';color:#08070d;text-decoration:none;'
                       'padding:13px 28px;border-radius:10px;font-size:13.5px;font-weight:600;">' + botao["label"] + '</a></p>')
    return (
        '<div style="font-family:Arial,Helvetica,sans-serif;max-width:480px;margin:0 auto;background:#060608;border-radius:20px;overflow:hidden;">'
        '<div style="background:linear-gradient(160deg,#0a0e2e 0%,#060810 100%);padding:28px 32px 22px;text-align:center;">'
        '<div style="font-family:Georgia,serif;font-size:22px;font-weight:bold;color:#fff;">atos<span style="color:#6db2ff;">.</span></div>'
        '</div>'
        '<div style="padding:28px 32px;color:#d4d4d8;font-size:14px;line-height:1.6;">'
        '<span style="display:inline-block;font-size:11.5px;font-weight:bold;border-radius:20px;padding:6px 14px;margin-bottom:16px;'
        'text-transform:uppercase;letter-spacing:0.4px;background:' + bg + ';color:' + cor + ';">' + status_label + '</span>'
        '<div style="font-size:18px;font-weight:bold;color:#fff;margin-bottom:10px;">' + titulo + '</div>'
        '<div style="font-size:13px;color:#8a90b8;margin-bottom:18px;">' + empresa_linha + '</div>'
        '<div style="background:rgba(255,255,255,0.04);border-radius:10px;padding:14px 16px;">' + info_html + '</div>'
        + botao_html +
        '</div>'
        '<div style="padding:20px 32px;border-top:1px solid rgba(255,255,255,0.06);text-align:center;">'
        '<div style="font-size:11px;color:#62666d;">contato@atos.net.br &middot; atos.net.br</div>'
        '</div>'
        '</div>'
    )


def _email_finalizado(p):
    corpo = corpo_status_cliente(p, "Finalizado", "Seu Processo foi Finalizado, em Anexo o Registro.")
    corpo_html = _email_status_html("finalizado", "Finalizado", "Processo Finalizado, em anexo o registro", _empresa_linha(p), protocolo=p.numero_protocolo or None)
    return corpo, corpo_html


def notificar_exigencia_cliente(db, p, origem="manual"):
    """Avisa o cliente por email quando o processo entra em exigencia.
    origem="autonoma": deteccao pela consulta autonoma (scraper, ver
    atualizar_status.py/aplicar_classificacao). origem="manual": registro
    manual pelo admin via POST /processos/{id}/exigencia.
    Sem PDF da exigencia + origem manual: nao dispara e-mail (comportamento ja
    existente preservado - nao e uma das variantes especificadas para o cliente,
    so o admin e alertado nesse caso pelo alerta administrativo separado)."""
    if not p.arquivo_exigencia and origem != "autonoma":
        return
    try:
        if p.arquivo_exigencia:
            corpo = corpo_status_cliente(p, "Exigencia", "Segue em anexo o documento da exigencia.")
            corpo_html = _email_status_html("exigencia", "Exigência", "Processo em Exigência", _empresa_linha(p),
                                             protocolo=p.numero_protocolo or None, nota_tipo="anexo", nota_texto="Exigência em Anexo")
            cam = os.path.join(UPLOADS_DIR, p.arquivo_exigencia)
            for em in emails_do_grupo(db, p.grupo_id):
                enviar_email_anexo(em, "Exigencia no seu processo - " + (p.empresa or ""), corpo, cam, p.arquivo_exigencia, corpo_html=corpo_html)
        else:
            corpo = corpo_status_cliente(p, "Exigencia", "Aguardando a Junta Comercial disponibilizar a exigencia.")
            corpo_html = _email_status_html("exigencia", "Exigência", "Processo em Exigência", _empresa_linha(p),
                                             protocolo=p.numero_protocolo or None, nota_tipo="aguardando",
                                             nota_texto="Aguardando a Junta Comercial disponibilizar a exigência.")
            for em in emails_do_grupo(db, p.grupo_id):
                enviar_email(em, "Exigencia no seu processo - " + (p.empresa or ""), corpo, corpo_html)
    except Exception as e:
        print("Erro ao notificar exigencia ao cliente:", e)


def notificar_tramitacao_cliente(db, p, status_antes):
    """Avisa o cliente por email quando o processo entra em tramitacao.
    Se so tiver o numero do protocolo, notifica so com o numero no corpo (sem anexo).
    Se tiver o pdf do protocolo, notifica com numero + anexo + aviso de protocolo em anexo.
    Compartilhada entre PATCH manual, upload de arquivo (Trocar/Salvar) e o Mane via
    Telegram, para nunca faltar aviso dependendo de qual caminho foi usado."""
    if (status_antes or "").lower() == "tramitacao" or (p.status or "").lower() != "tramitacao":
        return
    tem_numero = bool((p.numero_protocolo or "").strip())
    tem_pdf = bool((p.arquivo_protocolo or "").strip())
    if not tem_numero and not tem_pdf:
        print("Tramitacao sem email - falta numero e pdf.")
        return
    try:
        if tem_pdf:
            corpo = corpo_status_cliente(p, "Tramitacao", "Aguardando analise da Junta Comercial. Protocolo em anexo.")
            corpo_html = _email_status_html("tramitacao", "Tramitação", "Documento Protocolado", _empresa_linha(p),
                                             protocolo=p.numero_protocolo or None, nota_tipo="anexo", nota_texto="Protocolo em Anexo")
            cam = os.path.join(UPLOADS_DIR, p.arquivo_protocolo)
            for em in emails_do_grupo(db, p.grupo_id):
                enviar_email_anexo(em, "Atualizacao do seu processo - " + (p.empresa or ""), corpo, cam, p.arquivo_protocolo, corpo_html=corpo_html)
        else:
            corpo = corpo_status_cliente(p, "Tramitacao", "Aguardando analise da Junta Comercial.")
            corpo_html = _email_status_html("tramitacao", "Tramitação", "Documento Protocolado", _empresa_linha(p),
                                             protocolo=p.numero_protocolo or None)
            for em in emails_do_grupo(db, p.grupo_id):
                enviar_email(em, "Atualizacao do seu processo - " + (p.empresa or ""), corpo, corpo_html)
    except Exception as e:
        print("Erro ao notificar tramitacao:", e)


def vincular_fluxo_do_dia(db, processo, grupo_id):
    """Abre (se ainda nao existir) e vincula o processo ao Fluxo do dia do grupo,
    quando o grupo bate mais de 5 protocolos no mesmo dia. Chamar sempre logo apos
    setar processo.grupo_id, em todos os pontos que criam processo - mesmo padrao
    de nao deixar caminho de fora usado em notificar_tramitacao_cliente. Requer que
    o processo ja tenha sido db.add() + db.flush() antes desta chamada, para contar
    no total de hoje. Roda em SAVEPOINT (begin_nested) e nunca quebra o fluxo
    principal de criacao de processo: se falhar, so desfaz o que foi feito aqui
    dentro, sem afetar o processo ja pendente na transacao externa - mesmo
    espirito de robustez de registrar_evento."""
    if not grupo_id:
        return
    try:
        with db.begin_nested():
            hoje = date.today()
            fluxo = db.query(Fluxo).filter(Fluxo.grupo_id == grupo_id, Fluxo.data == hoje).first()
            total_hoje = db.query(Processo).filter(
                Processo.grupo_id == grupo_id,
                func.date(Processo.criado_em) == hoje
            ).count()

            if not fluxo and total_hoje > 5:
                fluxo = Fluxo(id=str(uuid.uuid4()), grupo_id=grupo_id, data=hoje, total_processos=total_hoje)
                db.add(fluxo)
                db.flush()

            if fluxo:
                processo.fluxo_id = fluxo.id
                fluxo.total_processos = total_hoje
    except Exception as e:
        print("Erro ao vincular fluxo do dia:", e)


def registrar_evento(db, processo, tipo, descricao, usuario=None):
    """Registra um evento pro feed de Atividade recente. Nao commita por conta
    propria - chamar sempre ANTES do commit() da operacao principal, para que o
    evento entre na mesma transacao (senao fica pendente e se perde no db.close()
    do get_db, que nao commita). Nunca deve quebrar o fluxo principal.

    usuario=None quando o evento e gerado por automacao sem usuario logado (ex:
    criacao automatica de processo de transferencia de sede) - a timeline mostra
    "Sistema (automação)" nesse caso (ver nome_usuario())."""
    try:
        evento = Evento(
            id=str(uuid.uuid4()), processo_id=processo.id,
            grupo_id=processo.grupo_id, tipo=tipo, descricao=descricao,
            usuario_login=getattr(usuario, "login", None) if usuario else None,
            usuario_nome=nome_usuario(usuario) if usuario else None,
            usuario_papel=(getattr(usuario, "papel", None) or ("admin" if getattr(usuario, "is_admin", False) else None)) if usuario else None,
        )
        db.add(evento)
    except Exception:
        pass


def rodape_atos():
    return (
        '<div style="border-top:1px solid #eef1f5;padding:18px 24px;background:#f7f9fc;">'
        '<div style="font-size:26px;font-weight:bold;color:#111111;letter-spacing:-1px;line-height:1;">atos<span style="color:#2d6cdf;">.</span></div>'
        '<div style="font-size:11px;color:#5a7088;letter-spacing:1px;margin-top:2px;">Gestao Societaria</div>'
        '<div style="font-size:11px;color:#9aa4b2;margin-top:8px;">contato@atos.net.br &middot; atos.net.br</div>'
        '</div>'
    )

def _badge_status(status_label):
    cores = {"Aberto": ("#eceae2", "#6b6c66"), "Tramitacao": ("#f0e0cb", "#8a5818"), "Exigencia": ("#f0dcd5", "#a8492a"), "Deferido": ("#d5e3df", "#15803d"), "Finalizado": ("#cfe8d8", "#15803d")}
    bg, cor = cores.get(status_label, ("#e6f1fb", "#185fa5"))
    return '<span style="display:inline-block;background:' + bg + ';color:' + cor + ';font-size:12px;font-weight:bold;padding:5px 14px;border-radius:20px;">' + status_label + '</span>'

def envolver_html(corpo_texto, titulo="Atualizacao do seu processo"):
    linhas = corpo_texto.split(chr(10))
    corpo_p = ""
    for ln in linhas:
        if ln.strip() == "":
            continue
        corpo_p = corpo_p + '<div style="font-size:14px;color:#445;line-height:1.65;margin-bottom:4px;">' + ln + '</div>'
    return (
        '<div style="font-family:Arial,Helvetica,sans-serif;max-width:520px;margin:0 auto;background:#ffffff;border:1px solid #e6ebf2;border-radius:12px;overflow:hidden;">'
        '<div style="height:5px;background:linear-gradient(90deg,#2563eb,#2dd4bf);"></div>'
        '<div style="padding:28px 24px;">'
        '<div style="font-size:19px;font-weight:bold;color:#1a2330;margin-bottom:14px;">' + titulo + '</div>'
        + corpo_p +
        '</div>'
        + rodape_atos() +
        '</div>'
    )


def _email_codigo_2fa_html(codigo):
    """Template do email de codigo de verificacao (2FA/login) - replica
    fielmente docs/email_codigo_v2.html (layout minimalista aprovado): fundo
    unico escuro, sem blocos separados, sem glow/blur/backdrop-filter (nao
    renderizam em clientes de email), cores solidas."""
    codigo_espacado = codigo[:3] + " " + codigo[3:]
    return (
        '<div style="font-family:-apple-system,\'Segoe UI\',Arial,sans-serif;max-width:440px;margin:0 auto;background:#0a0a0d;border-radius:16px;overflow:hidden;">'
        '<div style="padding:36px 32px 8px;text-align:center;">'
        '<div style="font-family:Georgia,serif;font-size:20px;font-weight:bold;color:#fff;">atos<span style="color:#6db2ff;">.</span></div>'
        '</div>'
        '<div style="padding:8px 32px 36px;color:#d4d4d8;font-size:14px;line-height:1.6;text-align:center;">'
        '<div style="font-size:15px;color:#8a90b8;margin-bottom:28px;font-weight:400;">Seu código de acesso</div>'
        '<div style="font-family:\'Courier New\',monospace;font-size:36px;font-weight:600;color:#fff;letter-spacing:6px;margin-bottom:20px;">' + codigo_espacado + '</div>'
        '<div style="font-size:12px;color:#62666d;">Válido por 10 minutos. Se não foi você, ignore este email.</div>'
        '</div>'
        '<div style="padding:18px 32px;text-align:center;">'
        '<div style="font-size:11px;color:#4a4a4e;"><a href="https://atos.net.br" style="color:#62666d;text-decoration:none;">atos.net.br</a></div>'
        '</div>'
        '</div>'
    )


def _disparar_convites(nome, link, emails):
    corpo = (
        "Olá!\n\n"
        "Você foi cadastrado para acessar o sistema Atos - Gestão Societária, no grupo " + nome + ".\n\n"
        "Para criar seu usuário e senha de acesso, clique no link abaixo:\n"
        + link + "\n\n"
        "Após criar seu acesso, você poderá acompanhar seus processos pelo endereço " + BASE_URL_SISTEMA + ".\n\n"
        "Atenciosamente,\nEquipe Atos"
    )
    corpo_html = (
        '<div style="font-family:Arial,Helvetica,sans-serif;max-width:480px;margin:0 auto;background:#060608;border-radius:20px;overflow:hidden;">'
        '<div style="background:linear-gradient(160deg,#0a0e2e 0%,#060810 100%);padding:36px 32px 28px;text-align:center;">'
        '<div style="font-family:Georgia,serif;font-size:26px;font-weight:bold;color:#fff;">atos<span style="color:#6db2ff;">.</span></div>'
        '<div style="font-size:11px;color:#9aa8d8;margin-top:4px;letter-spacing:0.5px;">GESTÃO SOCIETÁRIA</div>'
        '</div>'
        '<div style="padding:32px;color:#d4d4d8;font-size:14px;line-height:1.6;">'
        '<div style="font-family:Arial,Helvetica,sans-serif;font-size:19px;font-weight:bold;color:#fff;margin-bottom:14px;">Você foi cadastrado no Atos</div>'
        '<p style="margin:0 0 4px;">Você foi cadastrado para acessar o sistema Atos &mdash; Gestão Societária, no grupo:</p>'
        '<div style="display:inline-block;font-size:12px;color:#8ec2ff;background:rgba(77,148,255,0.1);border:1px solid rgba(77,148,255,0.3);border-radius:20px;padding:5px 14px;margin:14px 0 24px;">' + nome + '</div>'
        '<p>Clique no botão abaixo para criar seu login e senha de acesso.</p>'
        '<p style="text-align:center;margin-top:28px;"><a href="' + link + '" style="display:inline-block;background:linear-gradient(135deg,#4d94ff,#8c5aff);color:#ffffff;text-decoration:none;padding:14px 32px;border-radius:10px;font-size:14px;font-weight:600;">Criar meu acesso</a></p>'
        '<p style="font-size:12px;color:#62666d;margin-top:20px;">Ou copie e cole este endereço no navegador:<br><a href="' + link + '" style="color:#6db2ff;">' + link + '</a></p>'
        '</div>'
        '<div style="padding:24px 32px;border-top:1px solid rgba(255,255,255,0.06);text-align:center;">'
        '<div style="font-size:11px;color:#62666d;">contato@atos.net.br &middot; atos.net.br</div>'
        '</div>'
        '</div>'
    )
    for email in emails:
        enviar_email(email, "Acesso ao sistema Atos - " + nome, corpo, corpo_html)


app = FastAPI(title="Atos API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://atos.net.br", "https://www.atos.net.br"],
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOADS_DIR = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOADS_DIR, exist_ok=True)

# app.mount("/uploads", StaticFiles(directory=UPLOADS_DIR), name="uploads")  # desativado: arquivos agora so via /download protegido

GEMINI_KEY = os.getenv("GEMINI_KEY")
EMAIL_ADMIN = os.getenv("ADMIN_EMAIL")


def emails_admin(db):
    """Destinatarios dos alertas administrativos automaticos (exigencia, atraso,
    processo com campo incompleto, etc.): o admin fixo (EMAIL_ADMIN) mais todo
    usuario com papel 'operador' que tenha e-mail cadastrado. Centralizado aqui
    pra qualquer alerta admin - em main.py ou nos scripts de automacao que
    importam esta funcao - usar a mesma lista, sem duplicar quem recebe. Novo
    operador criado via /usuarios/operador entra automaticamente, sem precisar
    mudar codigo em nenhum ponto de disparo."""
    emails = [EMAIL_ADMIN] if EMAIL_ADMIN else []
    operadores = db.query(Usuario).filter(
        Usuario.papel == "operador",
        Usuario.email.isnot(None),
        Usuario.email != "",
    ).all()
    for u in operadores:
        if u.email and u.email not in emails:
            emails.append(u.email)
    return emails


DEEPSEEK_KEY = os.getenv("DEEPSEEK_KEY")
client = OpenAI(api_key=DEEPSEEK_KEY, base_url="https://api.deepseek.com")

INFOSIMPLES_TOKEN = os.getenv("INFOSIMPLES_TOKEN")
INFOSIMPLES_CPF = os.getenv("INFOSIMPLES_CPF")
INFOSIMPLES_SENHA_NFP = os.getenv("INFOSIMPLES_SENHA_NFP")
import sys as _sys
_sys.path.insert(0, os.path.join(BASE_DIR, "..", "automacao"))
from jucesp_infosimples import baixar_certidao_simplificada

CONHECIMENTO_FILE = r"D:\Mane\dados\conhecimento_registro.json"
def carregar_conhecimento():
    if os.path.exists(CONHECIMENTO_FILE):
        with open(CONHECIMENTO_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

CONHECIMENTO = carregar_conhecimento()

criar_banco()

# ============================================================
# ANALISAR ATA COM IA
# ============================================================
_CAMPOS_VAZIOS_ATA = {
    "empresa": "", "cnpj": "", "nire": "", "uf": "", "uf_destino_transferencia": "",
    "tipo_sociedade": "", "tipo_ato": "", "identificador_ato": "", "data_ata": "",
    "hora_ata": "", "email_cliente": "", "eventos": [], "requer_cpl": False,
    "checklist": [], "observacoes": "",
}


def analisar_ata_ia(texto_ata: str) -> dict:
    """PRIORIDADE MAXIMA: esta funcao NUNCA pode lancar excecao nem bloquear a
    criacao do processo. Se a IA falhar (erro de rede/API) ou devolver algo
    que nao seja JSON valido (texto de origem ruim/incompleto costuma causar
    isso), devolve os campos vazios em vez de propagar o erro - quem chama
    (criar_processo) ja trata campo vazio marcando o processo pra revisao
    manual, sem nunca bloquear a insercao."""
    conhecimento = json.dumps(CONHECIMENTO, ensure_ascii=False)[:3000]
    prompt = f"""Analise esta ata/documento e extraia as informações no formato JSON exato abaixo.

CONHECIMENTO BASE:
{conhecimento}

DOCUMENTO:
{texto_ata[:4000]}

REGRA IMPORTANTE PARA UF: Identifique a UF (sigla do estado, 2 letras) da sede da sociedade. Em alteracoes contratuais de sociedades limitadas, a UF aparece no campo de qualificacao da sociedade, no padrao Cidade/UF (exemplo: 'Rio de Janeiro/RJ' significa UF=RJ; 'Sao Paulo/SP' significa UF=SP). Procure a cidade seguida de barra e a sigla do estado no endereco da sede. Retorne so a sigla de 2 letras maiuscula.

REGRA IMPORTANTE PARA IDENTIFICAR A TITULAR DO ATO (campos empresa/cnpj/nire): o documento pode citar varias pessoas juridicas, mas so uma delas e a TITULAR do processo - aquela sobre quem recai a deliberacao do ato (o "RESOLVE"), normalmente identificada no CABECALHO do documento (nome em destaque no topo, com CNPJ e NIRE logo abaixo do titulo). Nao confunda a titular com quem apenas REPRESENTA, ASSINA EM NOME DE, ou E PARTE DELA na qualidade de procurador(a), representante legal, socio(a), consorciada ou EMPRESA LIDER - essas entidades podem aparecer primeiro no texto e ate assinar o documento, mas NAO sao a titular do ato, a menos que o proprio ato trate diretamente delas (e nao da entidade que representam).

CASO ESPECIFICO - ATOS DE CONSORCIO (Ato de Empresa Lider, Ato de Consorciada, etc.): o padrao tipico e "A [EMPRESA X], (...) na qualidade de Empresa Lider do CONSORCIO [NOME], (...) inscrito no CNPJ sob o no [CNPJ DO CONSORCIO] ('Consorcio'), neste ato representada por seu representante legal (...) RESOLVE: [ato sobre o Consorcio]". Nesses casos, a titular do ato e SEMPRE o CONSORCIO - use o nome, CNPJ e NIRE do CONSORCIO, nunca os da empresa lider/representante, mesmo que ela apareca primeiro no texto e assine o documento.

Exemplo de acerto (few-shot): um documento chamado "Ato da Empresa Lider do Consorcio sobre a Baixa do CNPJ" comeca dizendo "A 3R PETROLEUM OFFSHORE S.A., CNPJ 02.857.854/0001-14, na qualidade de Empresa Lider do CONSORCIO POT-M-475, CNPJ 20.045.684/0001-55, NIRE 33.5.0003193-0 ('Consorcio'), RESOLVE: autorizar a baixa do CNPJ do Consorcio...". Mesmo a 3R Petroleum aparecendo primeiro e assinando o documento, a resposta CORRETA e: empresa="CONSORCIO POT-M-475", cnpj="20.045.684/0001-55", nire="33.5.0003193-0" - NAO os dados da 3R Petroleum, que e so a representante/empresa lider, nao a titular do ato.

IMPORTANTE - O TEXTO ACIMA PODE ESTAR INCOMPLETO, COM RUIDO DE OCR OU PARCIALMENTE ILEGIVEL (documento escaneado de baixa qualidade). Mesmo assim:
- Faca sempre o seu MELHOR ESFORCO pra extrair o que for possivel identificar com confianca.
- NUNCA se recuse a responder e NUNCA retorne uma mensagem de erro em vez do JSON - mesmo que o documento esteja quase todo ilegivel, responda com a estrutura JSON completa.
- Para qualquer campo que voce nao conseguir identificar com confianca no texto disponivel, deixe-o vazio ("" ou [] ou false conforme o tipo) em vez de adivinhar ou recusar. E preferivel um campo vazio (revisado manualmente depois) do que um campo errado ou uma resposta fora do formato.

Retorne APENAS um JSON válido com esta estrutura exata:
{{
  "empresa": "nome completo da empresa",
  "cnpj": "XX.XXX.XXX/XXXX-XX",
  "nire": "número NIRE se encontrado",
  "uf": "sigla de 2 letras do estado da sede, ex RJ ou SP",
  "uf_destino_transferencia": "APENAS se a ata tratar de TRANSFERENCIA DE SEDE para outro Estado (mudanca de endereco da sede social de um Estado para outro, nao mudanca de endereco dentro do mesmo Estado): informe a sigla de 2 letras do Estado de DESTINO. Caso contrario deixe vazio.",
  "tipo_sociedade": "SA ou LTDA",
  "tipo_ato": "AGO, AGE, AGOE, RCA, ALTERACAO_CONTRATUAL, ARS etc",
  "identificador_ato": "ex: RCA 25/05/2026, 39ª Alteração Contratual, AGE 10/05/2026",
  "data_ata": "DD/MM/AAAA",
  "hora_ata": "HH:MM ou vazio",
  "email_cliente": "",
  "eventos": ["lista", "de", "eventos", "identificados"],
  "requer_cpl": true ou false,
  "checklist": ["lista", "de", "documentos", "necessários"],
  "observacoes": "alertas importantes"
}}"""

    try:
        resposta = client.chat.completions.create(
            model="deepseek-chat",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=1500,
            temperature=0.1
        )
        texto = resposta.choices[0].message.content
        texto_limpo = texto.replace("```json", "").replace("```", "").strip()
        dados = json.loads(texto_limpo)
    except Exception as e:
        print("analisar_ata_ia falhou (IA indisponivel ou resposta invalida) - devolvendo campos vazios pra nao bloquear:", str(e)[:200])
        return dict(_CAMPOS_VAZIOS_ATA)

    # Fallback: se a UF nao foi identificada pelo endereco da ata, infere pelo prefixo do NIRE
    # 333/332 = RJ | 353/352 = SP | 292/293 = BA | 262/263 = PE
    # (lista sera expandida com mais UFs futuramente)
    try:
        if not (dados.get("uf") or "").strip():
            nire_digitos = "".join(c for c in (dados.get("nire") or "") if c.isdigit())
            prefixo_nire = nire_digitos[:3]
            if prefixo_nire in ("333", "332"):
                dados["uf"] = "RJ"
            elif prefixo_nire in ("353", "352"):
                dados["uf"] = "SP"
            elif prefixo_nire in ("292", "293"):
                dados["uf"] = "BA"
            elif prefixo_nire in ("262", "263"):
                dados["uf"] = "PE"
    except Exception as e:
        print("Fallback de UF por NIRE falhou (ignorado, nao bloqueia):", str(e)[:150])

    return dados

# ============================================================
# ROTAS
# ============================================================
@app.get("/")
def root():
    return {"status": "Atos online"}


@app.post("/cadastro")
def cadastro(dados: dict, db: Session = Depends(get_db)):
    codigo_grupo = (dados.get("codigo_grupo") or "").strip()
    login = (dados.get("login") or "").strip()
    senha = dados.get("senha") or ""

    if not codigo_grupo or not login or not senha:
        raise HTTPException(status_code=400, detail="codigo_grupo, login e senha sao obrigatorios")
    if len(senha) < 6:
        raise HTTPException(status_code=400, detail="A senha deve ter pelo menos 6 caracteres")

    grupo = db.query(Grupo).filter(Grupo.codigo == codigo_grupo).first()
    if not grupo:
        raise HTTPException(status_code=400, detail="Codigo de grupo invalido")

    existente = db.query(Usuario).filter(Usuario.login == login).first()
    if existente:
        raise HTTPException(status_code=400, detail="Esse login ja esta em uso")

    senha_hash = bcrypt.hashpw(senha.encode()[:72], bcrypt.gensalt()).decode()
    novo = Usuario(
        id=str(uuid.uuid4()),
        login=login,
        senha_hash=senha_hash,
        grupo_id=grupo.id
    )
    db.add(novo)
    db.commit()
    return {"mensagem": "Usuario criado com sucesso", "login": login, "grupo": grupo.nome}


@app.post("/login")
def login(dados: dict, request: Request, db: Session = Depends(get_db)):
    ip = obter_ip(request) or "desconhecido"
    if not _checar_rate_login(ip):
        raise HTTPException(status_code=429, detail="Muitas tentativas. Tente novamente em alguns minutos.")
    login = (dados.get("login") or "").strip()
    senha = dados.get("senha") or ""

    if not login or not senha:
        raise HTTPException(status_code=400, detail="login e senha sao obrigatorios")

    usuario = db.query(Usuario).filter(Usuario.login == login).first()
    if not usuario or not bcrypt.checkpw(senha.encode()[:72], usuario.senha_hash.encode()):
        _registrar_falha_login(ip)
        raise HTTPException(status_code=401, detail="login ou senha invalidos")
    _limpar_falhas_login(ip)
    import random as _random
    codigo = "{:06d}".format(_random.randint(0, 999999))
    novo_cod = Codigo2FA(
        id=str(uuid.uuid4()),
        usuario_id=usuario.id,
        login=usuario.login,
        codigo=codigo,
        expira_em=datetime.now() + timedelta(minutes=10),
        usado=False,
    )
    db.add(novo_cod)
    db.commit()
    email_destino = email_do_usuario(db, usuario)
    if email_destino:
        try:
            corpo = "Seu codigo de acesso ao ATOS e: " + codigo + ". Valido por 10 minutos. Se voce nao tentou acessar, ignore este e-mail."
            enviar_email(email_destino, "Codigo de acesso ATOS", corpo, corpo_html=_email_codigo_2fa_html(codigo))
        except Exception as e:
            print("Erro ao enviar codigo 2FA:", e)
    return {"requer_2fa": True, "login": usuario.login, "mensagem": "Enviamos um codigo de acesso para o seu e-mail."}

@app.post("/login/verificar")
def login_verificar(dados: dict, request: Request, db: Session = Depends(get_db)):
    ip = obter_ip(request) or "desconhecido"
    login = (dados.get("login") or "").strip()
    codigo = (dados.get("codigo") or "").strip()
    if not login or not codigo:
        raise HTTPException(status_code=400, detail="login e codigo sao obrigatorios")
    usuario = db.query(Usuario).filter(Usuario.login == login).first()
    if not usuario:
        raise HTTPException(status_code=401, detail="usuario invalido")
    reg = db.query(Codigo2FA).filter(
        Codigo2FA.login == login,
        Codigo2FA.codigo == codigo,
        Codigo2FA.usado == False,
    ).order_by(Codigo2FA.criado_em.desc()).first()
    if not reg:
        raise HTTPException(status_code=401, detail="codigo invalido")
    if reg.expira_em < datetime.now():
        raise HTTPException(status_code=401, detail="codigo expirado, faca login novamente")
    reg.usado = True
    token = str(uuid.uuid4())
    usuario.token = token
    usuario.token_criado_em = datetime.now()
    db.commit()
    registrar_auditoria(db, usuario, "login", None, "acesso ao sistema (2FA)", ip)
    grupo = db.query(Grupo).filter(Grupo.id == usuario.grupo_id).first()
    return {
        "token": token,
        "login": usuario.login,
        "nome": getattr(usuario, "nome", None),
        "papel": getattr(usuario, "papel", None) or ("admin" if usuario.is_admin else "cliente"),
        "grupo_id": usuario.grupo_id,
        "grupo": grupo.nome if grupo else None,
        "is_admin": bool(usuario.is_admin),
    }




# ===== ANEXOS DO PROCESSO =====
def notificar_telegram(texto: str):
    """Envia um aviso ao ADM via Telegram. Retorna (chat_id, message_id) ou None."""
    try:
        import os, requests
        token = os.getenv("TELEGRAM_TOKEN")
        chat_id = os.getenv("TELEGRAM_CHAT_ID")
        if not token or not chat_id:
            return None
        r = requests.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            data={"chat_id": chat_id, "text": texto},
            timeout=5,
        )
        j = r.json()
        if j.get("ok"):
            return (str(chat_id), j["result"]["message_id"])
    except Exception:
        pass
    return None


def notificar_telegram_com_botoes(texto: str, botoes: list):
    """Igual notificar_telegram, mas com teclado inline (reply_markup).
    botoes: lista de {"texto": "...", "callback_data": "..."} - cada item
    vira um botao em linha propria. Usado pelo vigia normativo (Parte B)
    para as aprovacoes de Nivel 3 - primeira vez que o bot usa botao inline
    no projeto; bot.py precisa tratar callback_query (ver processar_callback)."""
    try:
        import os, requests, json as _json
        token = os.getenv("TELEGRAM_TOKEN")
        chat_id = os.getenv("TELEGRAM_CHAT_ID")
        if not token or not chat_id:
            return None
        teclado = {"inline_keyboard": [[{"text": b["texto"], "callback_data": b["callback_data"]}] for b in botoes]}
        r = requests.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            data={"chat_id": chat_id, "text": texto, "reply_markup": _json.dumps(teclado)},
            timeout=5,
        )
        j = r.json()
        if j.get("ok"):
            return (str(chat_id), j["result"]["message_id"])
    except Exception as e:
        print("Erro notificar_telegram_com_botoes:", str(e)[:150])
    return None

@app.post("/processos/{processo_id}/mensagens")
async def enviar_mensagem(processo_id: str, dados: str = Form(...), request: Request = None, x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    p = db.query(Processo).filter(Processo.id == processo_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")
    if not _tem_acesso_admin(usuario) and p.grupo_id != usuario.grupo_id:
        raise HTTPException(status_code=403, detail="Sem permissao para este processo")
    info = json.loads(dados)
    texto = (info.get("texto") or "").strip()
    if not texto:
        raise HTTPException(status_code=400, detail="Mensagem vazia")
    msg = MensagemProcesso(
        id=str(uuid.uuid4()),
        processo_id=processo_id,
        autor_login=usuario.login,
        autor_tipo=("admin" if usuario.is_admin else ("operador" if getattr(usuario, "papel", None) == "operador" else "cliente")),
        texto=texto,
        status_no_momento=p.status,
        tipo_ato_no_momento=p.tipo_ato,
    )
    db.add(msg)
    db.commit()
    _ip = obter_ip(request)
    registrar_auditoria(db, usuario, "mensagem_processo", processo_id, "", _ip)
    if not _tem_acesso_admin(usuario):
        _grupo = db.query(Grupo).filter(Grupo.id == p.grupo_id).first()
        _empresa = (_grupo.nome if _grupo else None) or p.empresa or "cliente"
        _ato = p.identificador_ato or p.tipo_ato or "processo"
        _preview = texto if len(texto) <= 500 else texto[:500] + "..."
        _aviso = f"O Cliente {_empresa}, no Processo: {_ato}, Usuario: {usuario.login}, fez uma pergunta:\n\n{_preview}"
        _res = notificar_telegram(_aviso)
        if _res:
            _cid, _mid = _res
            db.add(TelegramVinculo(id=str(uuid.uuid4()), telegram_message_id=_mid, chat_id=_cid, processo_id=processo_id))
            db.commit()
    return {"mensagem": "enviada", "id": msg.id}

@app.get("/processos/{processo_id}/mensagens")
async def listar_mensagens(processo_id: str, x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    p = db.query(Processo).filter(Processo.id == processo_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")
    if not _tem_acesso_admin(usuario) and p.grupo_id != usuario.grupo_id:
        raise HTTPException(status_code=403, detail="Sem permissao para este processo")
    msgs = db.query(MensagemProcesso).filter(MensagemProcesso.processo_id == processo_id).order_by(MensagemProcesso.criado_em.asc()).all()
    return [{"id": mm.id, "autor_login": mm.autor_login, "autor_tipo": mm.autor_tipo, "texto": mm.texto, "criado_em": mm.criado_em.isoformat() if mm.criado_em else None} for mm in msgs]

BASE_CONHECIMENTO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "docs", "base_conhecimento_atos_registros_juntas.md")


def _carregar_secoes_base_conhecimento():
    """Le o arquivo de base de conhecimento e devolve uma lista de secoes
    (### ou ## do markdown) com titulo e corpo. Recarrega do disco a cada
    chamada de proposito - o vigia normativo (Parte B) pode editar o
    arquivo entre uma pergunta e outra, e o Assistente sempre deve usar a
    versao mais atual."""
    try:
        with open(BASE_CONHECIMENTO_PATH, "r", encoding="utf-8") as f:
            texto = f.read()
    except Exception as e:
        print("Erro ao carregar base de conhecimento:", e)
        return []
    linhas = texto.split("\n")
    secoes = []
    titulo_atual = None
    corpo_atual = []
    for linha in linhas:
        if linha.startswith("### ") or linha.startswith("## "):
            if titulo_atual is not None:
                secoes.append({"titulo": titulo_atual, "corpo": "\n".join(corpo_atual).strip()})
            titulo_atual = linha.lstrip("#").strip()
            corpo_atual = [linha]
        else:
            corpo_atual.append(linha)
    if titulo_atual is not None:
        secoes.append({"titulo": titulo_atual, "corpo": "\n".join(corpo_atual).strip()})
    return secoes


def _normalizar_busca_texto(txt):
    import unicodedata
    t = unicodedata.normalize("NFKD", txt or "")
    t = "".join(c for c in t if not unicodedata.combining(c))
    return t.lower()


def _buscar_secoes_relevantes(termo_busca, top_n=2):
    """RAG simples por sobreposicao de palavras-chave (sem embeddings - a
    base tem ~40 secoes, e' suficiente e muito mais barato/rapido). Busca
    pelo tipo_ato/identificador_ato ja classificado no processo, nao manda
    o arquivo inteiro pro Gemini a cada pergunta."""
    import re
    secoes = _carregar_secoes_base_conhecimento()
    if not secoes:
        return []
    termo_norm = _normalizar_busca_texto(termo_busca)
    palavras_termo = set(w for w in re.split(r"\W+", termo_norm) if len(w) > 3)
    if not palavras_termo:
        return []
    pontuadas = []
    for s in secoes:
        titulo_norm = _normalizar_busca_texto(s["titulo"])
        corpo_norm = _normalizar_busca_texto(s["corpo"][:2000])
        pontos = 0
        for palavra in palavras_termo:
            if palavra in titulo_norm:
                pontos += 5
            elif palavra in corpo_norm:
                pontos += 1
        if pontos > 0:
            pontuadas.append((pontos, s))
    pontuadas.sort(key=lambda x: x[0], reverse=True)
    return [s for _, s in pontuadas[:top_n]]


def _gemini_assistente_atos(prompt):
    """Chama o Gemini pedindo JSON estruturado {resposta, confianca}. Tenta
    ate 3 vezes (1 tentativa original + 2 retries, com backoff de 1s/2s)
    antes de desistir - o iatos. precisa estar sempre disponivel pro
    cliente, uma falha ou timeout pontual da API nao pode virar erro na
    hora. Mesma convencao ja usada em _gemini_protocolo/_gemini_texto_documento
    (chamada REST direta via urllib, modelo gemini-flash-latest)."""
    import json as _json, urllib.request, time as _time
    if not GEMINI_KEY:
        return None
    ultimo_erro = None
    for tentativa in range(3):
        try:
            body = {"contents": [{"parts": [{"text": prompt}]}]}
            url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-flash-latest:generateContent?key=" + GEMINI_KEY
            req = urllib.request.Request(url, data=_json.dumps(body).encode(), headers={"Content-Type": "application/json"})
            resp = urllib.request.urlopen(req, timeout=40)
            data = _json.loads(resp.read().decode())
            txt = data["candidates"][0]["content"]["parts"][0]["text"].strip()
            txt = txt.replace("```json", "").replace("```", "").strip()
            resultado = _json.loads(txt)
            if resultado.get("resposta"):
                return resultado
            ultimo_erro = "resposta do Gemini sem campo 'resposta' valido"
        except Exception as e:
            ultimo_erro = str(e)[:200]
        if tentativa < 2:
            print(f"   [iatos.] tentativa {tentativa + 1}/3 falhou ({ultimo_erro}), retentando...")
            _time.sleep(1 + tentativa)
    print("   [iatos.] todas as 3 tentativas falharam:", ultimo_erro)
    return None


def _contexto_processo_assistente(p):
    partes = [
        "Empresa: " + (p.empresa or "-"),
        "CNPJ: " + (p.cnpj or "-"),
        "Tipo de sociedade: " + (p.tipo_sociedade or "-"),
        "Tipo de ato: " + (p.tipo_ato or "-"),
        "Identificador do ato: " + (p.identificador_ato or "-"),
        "UF/Junta: " + (p.uf or "-"),
        "Status atual: " + (p.status or "-"),
        "Data de insercao no sistema: " + (p.criado_em.strftime("%d/%m/%Y %H:%M") if p.criado_em else "-"),
        "Numero de protocolo: " + (p.numero_protocolo or "ainda nao protocolado"),
    ]
    if p.status_jucesp:
        partes.append("Ultimo status retornado pela Junta: " + p.status_jucesp)
    if p.texto_exigencia:
        partes.append("Texto da exigencia atual: " + p.texto_exigencia)
    return "\n".join(partes)


def _obter_texto_documento_processo(db, p):
    """Texto integral do documento/ata deste processo. Se ja foi salvo na
    insercao (fluxo normal, a partir desta mudanca), usa direto. Se for um
    processo antigo (inserido antes dessa coluna existir) e ainda tiver o
    PDF salvo, extrai na hora e PERSISTE de volta - assim so' precisa
    re-OCRizar uma vez por processo antigo, nao a cada pergunta."""
    if p.texto_documento_extraido:
        return p.texto_documento_extraido
    if not p.arquivo_ata:
        return ""
    caminho = os.path.join(UPLOADS_DIR, p.arquivo_ata)
    if not os.path.exists(caminho):
        return ""
    try:
        texto, _parcial = extrair_texto_pdf_em_camadas(caminho)
        if texto:
            p.texto_documento_extraido = texto
            db.commit()
        return texto or ""
    except Exception as e:
        print("   [iatos.] falha ao extrair texto do documento sob demanda:", str(e)[:200])
        return ""


# Cache em memoria do pacote de contexto completo por processo (texto integral
# do documento + dados estruturados + secao(oes) da base de conhecimento) -
# montado UMA VEZ (ao abrir o chat, ou na 1a pergunta se "abrir" nao foi
# chamado) e reutilizado em toda pergunta subsequente da mesma sessao, nunca
# remontado/decidido pergunta a pergunta pela IA. TTL curto so' pra nao ficar
# desatualizado numa sessao de chat muito longa (status pode mudar entre
# perguntas, ex: o cron rodou no meio da conversa).
_CONTEXTO_IATOS_CACHE = {}
_CONTEXTO_IATOS_TTL_SEGUNDOS = 60 * 60


def _montar_contexto_iatos(db, p, forcar=False):
    agora = datetime.now()
    cache = _CONTEXTO_IATOS_CACHE.get(p.id)
    if not forcar and cache and (agora - cache["montado_em"]).total_seconds() < _CONTEXTO_IATOS_TTL_SEGUNDOS:
        return cache

    texto_doc = _obter_texto_documento_processo(db, p)
    termo_busca = (p.tipo_ato or "") + " " + (p.identificador_ato or "")
    secoes = _buscar_secoes_relevantes(termo_busca, top_n=3)
    secoes_texto = "\n\n---\n\n".join(s["corpo"] for s in secoes) if secoes else "(nenhuma secao especifica encontrada na base para esse tipo de ato)"

    bloco = f"""CONTEXTO DO PROCESSO:
{_contexto_processo_assistente(p)}

TEXTO INTEGRAL DO DOCUMENTO/ATA DESTE PROCESSO ESPECIFICO:
{texto_doc[:12000] if texto_doc else "(texto do documento nao disponivel - nem salvo no processo, nem foi possivel extrair do arquivo)"}

BASE DE CONHECIMENTO NORMATIVA (secao(oes) relevante(s) ao tipo de ato deste processo):
{secoes_texto}"""

    entrada = {"texto": bloco, "montado_em": agora, "secoes_titulos": [s["titulo"] for s in secoes], "tinha_texto_doc": bool(texto_doc)}
    _CONTEXTO_IATOS_CACHE[p.id] = entrada
    return entrada


@app.post("/assistente/abrir")
async def assistente_abrir(dados: dict = Body(...), x_token: str = Header(None), db: Session = Depends(get_db)):
    """Chamado no exato momento em que o cliente clica no icone do iatos.,
    ANTES de qualquer pergunta - monta (ou remonta, forcado) o pacote de
    contexto completo do processo e deixa cacheado, pronto pra ser
    reutilizado em toda pergunta feita durante a sessao de chat."""
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    processo_id = dados.get("processo_id")
    if not processo_id:
        raise HTTPException(status_code=400, detail="processo_id e obrigatorio")
    p = db.query(Processo).filter(Processo.id == processo_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")
    if not _tem_acesso_admin(usuario) and p.grupo_id != usuario.grupo_id:
        raise HTTPException(status_code=403, detail="Sem permissao para este processo")
    entrada = _montar_contexto_iatos(db, p, forcar=True)
    return {"ok": True, "tem_texto_documento": entrada["tinha_texto_doc"], "secoes_usadas": entrada["secoes_titulos"]}


@app.post("/assistente/perguntar")
async def assistente_perguntar(dados: dict = Body(...), x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    processo_id = dados.get("processo_id")
    pergunta = (dados.get("pergunta") or "").strip()
    historico = dados.get("historico") or []
    if not processo_id or not pergunta:
        raise HTTPException(status_code=400, detail="processo_id e pergunta sao obrigatorios")
    p = db.query(Processo).filter(Processo.id == processo_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")
    if not _tem_acesso_admin(usuario) and p.grupo_id != usuario.grupo_id:
        raise HTTPException(status_code=403, detail="Sem permissao para este processo")

    # Mensagem unica, literal, pros dois motivos de escalonamento (falha tecnica do
    # Gemini mesmo apos os retries, OU baixa confianca de conteudo) - do ponto de
    # vista do cliente o resultado e' o mesmo (alguem da equipe assume), entao NUNCA
    # mostra um erro tecnico cru nem a resposta parcial de baixa confianca.
    MENSAGEM_FALLBACK_IATOS = "Sua dúvida é específica e por isso um Operador Atos vai entrar em contato. Obrigado."

    resultado = None
    motivo_escalonamento = None
    secoes_titulos = []
    try:
        # Reusa o contexto ja montado (ao abrir o chat, ou em pergunta anterior
        # da mesma sessao) - NUNCA decide pergunta a pergunta se busca o
        # documento/base; isso ja esta garantido e cacheado desde o "abrir".
        contexto = _montar_contexto_iatos(db, p)
        secoes_titulos = contexto["secoes_titulos"]

        historico_texto = ""
        if historico:
            linhas_hist = []
            for h in historico[-6:]:
                papel = "Cliente" if h.get("autor") == "usuario" else "iatos."
                linhas_hist.append(papel + ": " + str(h.get("texto") or ""))
            historico_texto = "\n".join(linhas_hist)

        prompt = f"""Voce e o "iatos.", assistente de IA que ajuda clientes do sistema ATOS (gestao de
registros em Juntas Comerciais) a entender o andamento e as implicacoes juridicas do ato
societario especifico deles. Responda em portugues, de forma direta e acessivel (o usuario
normalmente NAO e advogado).

Voce TEM ACESSO ao documento/ata completo deste processo especifico E a secao pertinente da
base de conhecimento normativa (ambos fornecidos abaixo) - CRUZE as duas fontes pra responder:
use o documento pra saber o que JA FOI DEFINIDO nesse caso especifico (fundamento legal citado,
publicacoes ja autorizadas, prazos ja fixados na propria ata, etc.) e a base de conhecimento pra
confirmar/completar com a regra normativa geral (proximos passos, codigo de evento, requisitos
de DBE, etc.). NUNCA invente informacao que nao esteja em nenhuma das duas fontes.

So' escale a resposta ao Operador Atos (classificando confianca como "baixa") se a pergunta
exigir informacao que nao esta em NENHUMA das duas fontes (nem no documento, nem na base), ou
envolver interpretacao juridica de nuance/caso concreto que a base nao resolve com seguranca.
Se o documento ou a base ja tem a resposta, responda com confianca "alta" ou "media" - nao
escale so' por precaucao quando a informacao esta disponivel.

{contexto["texto"]}

{"HISTORICO DA CONVERSA ATE AGORA:\n" + historico_texto if historico_texto else ""}

PERGUNTA DO CLIENTE:
{pergunta}

Responda APENAS com um JSON no formato exato:
{{"resposta": "sua resposta ao cliente, em texto corrido", "confianca": "alta" | "media" | "baixa"}}"""

        resultado = _gemini_assistente_atos(prompt)
    except Exception as e:
        # Qualquer excecao NAO tratada aqui (RAG, montagem de contexto, etc, nao so'
        # a chamada do Gemini em si) tambem cai no fallback - o iatos. tem que
        # responder sempre, nunca devolver um 500 cru pro cliente.
        print("   [iatos.] excecao inesperada montando resposta:", str(e)[:300])
        resultado = None

    if not resultado or not resultado.get("resposta"):
        resposta_final = MENSAGEM_FALLBACK_IATOS
        confianca = "baixa"
        motivo_escalonamento = "falha_tecnica"
    else:
        confianca = (resultado.get("confianca") or "media").lower()
        if confianca not in ("alta", "media", "baixa"):
            confianca = "media"
        if confianca == "baixa":
            resposta_final = MENSAGEM_FALLBACK_IATOS
            motivo_escalonamento = "baixa_confianca"
        else:
            resposta_final = resultado["resposta"]

    escalado = motivo_escalonamento is not None
    if escalado:
        _grupo = db.query(Grupo).filter(Grupo.id == p.grupo_id).first()
        _empresa_nome = (_grupo.nome if _grupo else None) or p.empresa or "cliente"
        _ato = p.identificador_ato or p.tipo_ato or "processo"
        motivo_label = (
            "FALHA TECNICA (Gemini indisponivel/erro mesmo apos retries, ou excecao no backend)"
            if motivo_escalonamento == "falha_tecnica"
            else "BAIXA CONFIANCA (base de conhecimento nao cobre com seguranca)"
        )
        resposta_ia_bruta = resultado.get("resposta") if resultado else None
        aviso = (
            f"[iatos. - escalonado: {motivo_label}]\n"
            f"Cliente: {_empresa_nome} | Processo: {_ato} | Usuario: {usuario.login}\n\n"
            f"Pergunta: {pergunta}\n\n"
            + (f"Resposta que a IA daria (nao mostrada ao cliente): {resposta_ia_bruta}\n\n" if resposta_ia_bruta else "")
            + "Cliente recebeu a mensagem padrao de encaminhamento a um operador."
        )
        notificar_telegram(aviso)

    conversa = AssistenteConversa(
        id=str(uuid.uuid4()),
        processo_id=processo_id,
        usuario_id=usuario.id,
        mensagem=pergunta,
        resposta=resposta_final,
        nivel_confianca=confianca,
        secao_usada=", ".join(secoes_titulos) if secoes_titulos else None,
        escalado_admin=escalado,
        motivo_escalonamento=motivo_escalonamento,
    )
    db.add(conversa)
    db.commit()

    return {"resposta": resposta_final, "confianca": confianca, "escalado": escalado}


@app.get("/processos/{processo_id}/assistente/historico")
async def assistente_historico(processo_id: str, x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    p = db.query(Processo).filter(Processo.id == processo_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")
    if not _tem_acesso_admin(usuario) and p.grupo_id != usuario.grupo_id:
        raise HTTPException(status_code=403, detail="Sem permissao para este processo")
    conversas = db.query(AssistenteConversa).filter(AssistenteConversa.processo_id == processo_id).order_by(AssistenteConversa.criado_em.asc()).all()
    return [{
        "id": c.id, "mensagem": c.mensagem, "resposta": c.resposta,
        "nivel_confianca": c.nivel_confianca, "escalado_admin": c.escalado_admin,
        "criado_em": c.criado_em.isoformat() if c.criado_em else None,
    } for c in conversas]


@app.post("/processos/{processo_id}/anexos")
async def enviar_anexo(processo_id: str, arquivo: UploadFile = File(...), descricao: str = Form(None), request: Request = None, x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    p = db.query(Processo).filter(Processo.id == processo_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")
    if not _tem_acesso_admin(usuario) and p.grupo_id != usuario.grupo_id:
        raise HTTPException(status_code=403, detail="Sem permissao para este processo")
    ext = os.path.splitext(arquivo.filename or "")[1].lower()
    EXT_PERMITIDAS = {".pdf", ".png", ".jpg", ".jpeg", ".xml", ".txt"}
    if ext not in EXT_PERMITIDAS:
        raise HTTPException(status_code=400, detail="Tipo de arquivo nao permitido para anexo.")
    conteudo = await arquivo.read()
    if len(conteudo) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Arquivo muito grande. Limite de 20 MB.")
    if len(conteudo) == 0:
        raise HTTPException(status_code=400, detail="Arquivo vazio.")
    anexo_id = str(uuid.uuid4())
    nome_arquivo = "anexo_" + anexo_id + ext
    caminho = os.path.join(UPLOADS_DIR, nome_arquivo)
    with open(caminho, "wb") as f:
        f.write(conteudo)
    novo = Anexo(id=anexo_id, processo_id=processo_id, arquivo=nome_arquivo, nome_original=(arquivo.filename or ""), descricao=(descricao or ""), enviado_por=usuario.login)
    db.add(novo)
    db.commit()
    _ip = obter_ip(request)
    registrar_auditoria(db, usuario, "anexo_upload", processo_id, "arquivo=" + (arquivo.filename or ""), _ip)
    return {"mensagem": "Anexo enviado", "id": anexo_id, "nome_original": arquivo.filename}

@app.get("/processos/{processo_id}/anexos")
def listar_anexos(processo_id: str, x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    p = db.query(Processo).filter(Processo.id == processo_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")
    if not _tem_acesso_admin(usuario) and p.grupo_id != usuario.grupo_id:
        raise HTTPException(status_code=403, detail="Sem permissao para este processo")
    anexos = db.query(Anexo).filter(Anexo.processo_id == processo_id).order_by(Anexo.criado_em).all()
    return [{"id": a.id, "nome_original": a.nome_original, "descricao": a.descricao, "enviado_por": a.enviado_por, "criado_em": str(a.criado_em)} for a in anexos]

@app.get("/anexos/{anexo_id}/download")
def baixar_anexo(anexo_id: str, request: Request = None, x_token: str = Header(None), db: Session = Depends(get_db)):
    from fastapi.responses import FileResponse
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    a = db.query(Anexo).filter(Anexo.id == anexo_id).first()
    if not a:
        raise HTTPException(status_code=404, detail="Anexo nao encontrado")
    p = db.query(Processo).filter(Processo.id == a.processo_id).first()
    if p and not _tem_acesso_admin(usuario) and p.grupo_id != usuario.grupo_id:
        raise HTTPException(status_code=403, detail="Sem permissao para este anexo")
    caminho = os.path.join(UPLOADS_DIR, a.arquivo)
    if not os.path.exists(caminho):
        raise HTTPException(status_code=404, detail="Arquivo nao encontrado no disco")
    _ip = obter_ip(request)
    registrar_auditoria(db, usuario, "anexo_download", a.processo_id, "anexo=" + (a.nome_original or ""), _ip)
    return FileResponse(caminho, filename=(a.nome_original or a.arquivo))

@app.delete("/anexos/{anexo_id}")
def excluir_anexo(anexo_id: str, request: Request = None, x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    a = db.query(Anexo).filter(Anexo.id == anexo_id).first()
    if not a:
        raise HTTPException(status_code=404, detail="Anexo nao encontrado")
    _ip = obter_ip(request)
    if not usuario.is_admin and a.enviado_por != usuario.login:
        registrar_auditoria(db, usuario, "anexo_excluir_negado", a.processo_id, "anexo=" + (a.nome_original or "") + " motivo=sem_permissao", _ip)
        raise HTTPException(status_code=403, detail="Apenas administrador ou quem enviou o anexo pode exclui-lo")
    caminho = os.path.join(UPLOADS_DIR, a.arquivo)
    try:
        if os.path.exists(caminho):
            os.remove(caminho)
    except Exception as e:
        print("erro ao remover anexo do disco:", e)
    proc_id = a.processo_id
    nome = a.nome_original
    db.delete(a)
    db.commit()
    registrar_auditoria(db, usuario, "anexo_excluir", proc_id, "anexo=" + (nome or ""), _ip)
    return {"mensagem": "Anexo removido"}

@app.post("/processos/{processo_id}/certidao-simplificada")
def emitir_certidao_simplificada(processo_id: str, request: Request = None, x_token: str = Header(None), db: Session = Depends(get_db)):
    """Emite a Certidao Simplificada da JUCESP (via Infosimples) pro NIRE do
    processo e anexa o PDF ao processo. NAO notifica o cliente automaticamente
    - so disponibiliza o documento no sistema (anexo), por decisao explicita:
    se isso deve entrar no e-mail automatico de "processo concluido" fica pra
    decidir depois."""
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    p = db.query(Processo).filter(Processo.id == processo_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")
    if not _tem_acesso_admin(usuario) and p.grupo_id != usuario.grupo_id:
        raise HTTPException(status_code=403, detail="Sem permissao para este processo")
    if not p.nire:
        raise HTTPException(status_code=400, detail="Processo sem NIRE cadastrado")
    if not all([INFOSIMPLES_TOKEN, INFOSIMPLES_CPF, INFOSIMPLES_SENHA_NFP]):
        raise HTTPException(status_code=503, detail="Credenciais da Infosimples nao configuradas")

    _ip = obter_ip(request)
    anexo_id = str(uuid.uuid4())
    nome_arquivo = "certidao_simplificada_" + anexo_id + ".pdf"
    caminho = os.path.join(UPLOADS_DIR, nome_arquivo)
    ok = baixar_certidao_simplificada(p.nire, INFOSIMPLES_TOKEN, INFOSIMPLES_CPF, INFOSIMPLES_SENHA_NFP, caminho)
    if not ok:
        registrar_auditoria(db, usuario, "certidao_simplificada_erro", processo_id, "nire=" + str(p.nire), _ip)
        raise HTTPException(status_code=502, detail="Nao foi possivel emitir a certidao simplificada (NIRE invalido, credencial invalida ou falha na Infosimples)")

    novo = Anexo(
        id=anexo_id,
        processo_id=processo_id,
        arquivo=nome_arquivo,
        nome_original="Certidao Simplificada JUCESP - " + (p.empresa or p.nire) + ".pdf",
        descricao="Certidao Simplificada emitida automaticamente via Infosimples",
        enviado_por=usuario.login,
    )
    db.add(novo)
    db.commit()
    registrar_auditoria(db, usuario, "certidao_simplificada_emitida", processo_id, "nire=" + str(p.nire), _ip)
    return {
        "mensagem": "Certidao Simplificada emitida e anexada ao processo",
        "anexo_id": anexo_id,
        "nome_original": novo.nome_original,
        "download_url": "/anexos/" + anexo_id + "/download",
    }

@app.get("/download/{processo_id}/{tipo}")
def download(processo_id: str, tipo: str, request: Request = None, x_token: str = Header(None), db: Session = Depends(get_db)):
    from fastapi.responses import FileResponse
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    p = db.query(Processo).filter(Processo.id == processo_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")
    if not _tem_acesso_admin(usuario) and p.grupo_id != usuario.grupo_id:
        raise HTTPException(status_code=403, detail="Acesso negado a este processo")
    campo_map = {"ata": p.arquivo_ata, "protocolo": p.arquivo_protocolo, "registro": p.arquivo_registro, "nd": p.arquivo_nd, "nf": p.arquivo_nf, "exigencia": p.arquivo_exigencia}
    if tipo not in campo_map:
        raise HTTPException(status_code=400, detail="Tipo invalido")
    nome_arquivo = campo_map[tipo]
    if not nome_arquivo:
        raise HTTPException(status_code=404, detail="Arquivo nao disponivel")
    caminho = os.path.join(UPLOADS_DIR, nome_arquivo)
    if not os.path.exists(caminho):
        raise HTTPException(status_code=404, detail="Arquivo nao encontrado no disco")
    _ip = obter_ip(request)
    registrar_auditoria(db, usuario, "download", processo_id, "tipo=" + str(tipo) + " arquivo=" + str(nome_arquivo), _ip)
    return FileResponse(caminho, filename=nome_arquivo)

@app.get("/processos")
def listar_processos(codigo_grupo: str = None, x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    # defer: o texto integral do documento (pode ser 10k+ caracteres, conteudo
    # sensivel de ata societaria) nunca precisa ir pro frontend na listagem -
    # so' e' lido server-side, sob demanda, pelo iatos. (ver _montar_contexto_iatos).
    query = db.query(Processo).options(defer(Processo.texto_documento_extraido))
    if _tem_acesso_admin(usuario):
        if codigo_grupo:
            grupo = db.query(Grupo).filter(Grupo.codigo == codigo_grupo).first()
            if grupo:
                query = query.filter(Processo.grupo_id == grupo.id)
    else:
        query = query.filter(Processo.grupo_id == usuario.grupo_id)
    from sqlalchemy import case
    processos = query.order_by(
        case((Processo.status == "finalizado", 1), else_=0),
        case((Processo.status == "finalizado", Processo.atualizado_em), else_=Processo.criado_em).desc()
    ).all()
    return processos

@app.get("/processos/pendentes")
async def listar_pendentes(x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    requer_acesso_admin(usuario)
    ps = db.query(Processo).filter(Processo.confirmacao_pendente == True).all()
    return [{"id": p.id, "empresa": p.empresa, "tipo_ato": p.tipo_ato, "tipo_ato_sugerido": p.tipo_ato_sugerido, "identificador_ato": p.identificador_ato, "data_ata": p.data_ata} for p in ps]

@app.get("/processos/checar-duplicidade")
async def checar_duplicidade(empresa: str = "", tipo_ato: str = "", data_ata: str = "", hora_ata: str = "", identificador_ato: str = "", x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    q = db.query(Processo)
    if not _tem_acesso_admin(usuario):
        q = q.filter(Processo.grupo_id == usuario.grupo_id)
    alvo = (_norm(empresa), _norm(tipo_ato), _norm(data_ata), _norm(hora_ata), _norm(identificador_ato))
    for p in q.all():
        atual = (_norm(p.empresa), _norm(p.tipo_ato), _norm(p.data_ata), _norm(p.hora_ata), _norm(p.identificador_ato))
        if atual == alvo and any(alvo):
            return {"duplicado": True, "processo_id": p.id, "empresa": p.empresa, "identificador_ato": p.identificador_ato}
    return {"duplicado": False}

# IMPORTANTE: as duas rotas acima (pendentes, checar-duplicidade) tem que vir
# ANTES de /processos/{processo_id} - FastAPI casa rotas por ordem de
# registro, nao por especificidade. Como {processo_id} aceita qualquer string,
# se ele viesse primeiro "engoliria" essas duas rotas literais (bug real ja
# visto em producao: GET /processos/pendentes retornava 404 do backend).
@app.get("/processos/{processo_id}")
def obter_processo(processo_id: str, request: Request = None, x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    p = db.query(Processo).options(defer(Processo.texto_documento_extraido)).filter(Processo.id == processo_id).first()
    if p and not _tem_acesso_admin(usuario) and p.grupo_id != usuario.grupo_id:
        raise HTTPException(status_code=403, detail="Sem permissao para este processo")
    if not p:
        raise HTTPException(status_code=404, detail="Processo não encontrado")
    _ip = obter_ip(request)
    registrar_auditoria(db, usuario, "visualizar", processo_id, "empresa=" + str(p.empresa), _ip)
    return p

@app.get("/fluxo/ativo")
def fluxo_ativo(codigo_grupo: str = None, x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    grupo_id_filtro = None
    if _tem_acesso_admin(usuario):
        if codigo_grupo:
            grupo = db.query(Grupo).filter(Grupo.codigo == codigo_grupo).first()
            if not grupo:
                return None  # codigo_grupo invalido -> nao ha fluxo pra retornar, nao cai na busca sem filtro
            grupo_id_filtro = grupo.id
    else:
        grupo_id_filtro = usuario.grupo_id

    hoje = date.today()
    query = db.query(Fluxo).filter(Fluxo.data == hoje)
    if grupo_id_filtro:
        query = query.filter(Fluxo.grupo_id == grupo_id_filtro)
    fluxos = query.all()

    resultado = []
    for f in fluxos:
        processos = db.query(Processo).filter(Processo.fluxo_id == f.id).all()
        pendentes = [p for p in processos if p.status != "finalizado"]
        if not pendentes:
            continue  # todos finalizaram -> nao retorna, card some
        confirmados = len([p for p in processos if p.status in ("deferido", "finalizado")])
        resultado.append({
            "grupo_id": f.grupo_id,
            "data": f.data.isoformat(),
            "total": len(processos),
            "confirmados": confirmados,
            "em_tramitacao": len([p for p in processos if p.status == "tramitacao"]),
        })

    if _tem_acesso_admin(usuario) and not codigo_grupo:
        return resultado
    return resultado[0] if resultado else None


@app.get("/eventos/recentes")
def eventos_recentes(codigo_grupo: str = None, limit: int = 10, x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    query = db.query(Evento).order_by(Evento.criado_em.desc())
    if _tem_acesso_admin(usuario):
        if codigo_grupo:
            grupo = db.query(Grupo).filter(Grupo.codigo == codigo_grupo).first()
            if grupo:
                query = query.filter(Evento.grupo_id == grupo.id)
    else:
        query = query.filter(Evento.grupo_id == usuario.grupo_id)
    eventos = query.limit(limit).all()
    # nome/papel do autor so vao pra quem tem acesso a tela administrativa - o
    # cliente continua vendo so a descricao, sem nome de quem da equipe agiu.
    mostrar_autor = _tem_acesso_admin(usuario)
    return [{
        "tipo": e.tipo,
        "descricao": e.descricao,
        "processo_id": e.processo_id,
        "criado_em": e.criado_em.isoformat(),
        "autor_nome": (e.usuario_nome or "Sistema (automação)") if mostrar_autor else None,
        "autor_papel": e.usuario_papel if mostrar_autor else None,
    } for e in eventos]

# ===== PARTE 2: deteccao automatica do documento principal =====
TIPOS_PRINCIPAIS = {
    "Contrato Social": ["contrato social"],
    "Alteracao Contratual": ["alteracao do contrato social", "alteracao contratual", "alteração do contrato social", "alteração contratual", "alteracao e consolidacao do contrato social", "alteração e consolidação do contrato social", "alteracao e consolidacao de contrato social", "alteração e consolidação de contrato social", "consolidacao do contrato social", "consolidação do contrato social"],
    "Ata de Reuniao/Assembleia de Socios": ["ata de reuniao de socios", "ata de assembleia de socios", "reuniao de socios", "ata de reunião de sócios", "ata de assembleia de sócios", "reunião de sócios"],
    "Distrato/Dissolucao/Liquidacao": ["distrato", "dissolucao", "liquidacao", "dissolução", "liquidação"],
    "Estatuto Social": ["estatuto social"],
    "Ata de Assembleia Geral de Constituicao": ["assembleia geral de constituicao", "assembleia geral de constituição"],
    "Ata de AGO": ["assembleia geral ordinaria", "assembleia geral ordinária"],
    "Ata de AGE": ["assembleia geral extraordinaria", "assembleia geral extraordinária"],
    "Ata de Reuniao do Conselho de Administracao": ["reuniao do conselho", "conselho de administra", "reunião do conselho"],
    "Ata de Reuniao de Diretoria": ["reuniao de diretoria", "reunião de diretoria"],
    "Escritura de Emissao de Debentures": ["escritura de emissao de debentures", "emissao de debentures", "escritura de emissão de debêntures", "emissão de debêntures"],
    "Boletim/Lista/Carta de Subscricao": ["boletim de subscricao", "lista de subscricao", "carta de subscricao", "boletim de subscrição", "lista de subscrição", "carta de subscrição"],
    "Ata de Assembleia Geral": ["ata de assembleia geral", "ata da assembleia geral"],
}
MARCADORES_ANEXO = [
    "requerimento", "ficha de cadastro nacional", "consulta de viabilidade",
    "documento basico de entrada", "documento básico de entrada",
    "procuracao", "procuração", "declaracao de desimpedimento", "declaração de desimpedimento",
    "darf", "gare", "comprovante de pagamento", "comprovante", "certidao", "certidão",
    "balanco patrimonial", "balanço patrimonial", "sped", "prospecto",
    "diario oficial", "diário oficial",
    "carteira de identidade", "documento de identidade", "doc identidade", "identidade",
    "lista de presenca", "lista de presenca de socios", "lista de presenca de acionistas", "registro geral", "cnh", "carteira nacional de habilitacao", "carteira nacional de habilitação", "habilitacao", "habilitação",
]
EXT_IMAGEM = {".jpg", ".jpeg", ".png"}

def _gemini_texto_documento(caminho_pdf):
    import base64, json, urllib.request
    if not GEMINI_KEY:
        return None
    try:
        with open(caminho_pdf, "rb") as f:
            pdf_b64 = base64.b64encode(f.read()).decode()
        prompt = ("Transcreva todo o texto visivel deste documento (ata, certidao ou comprovante de registro "
                  "de Junta Comercial), incluindo carimbos, selos e textos de certificacao de registro/arquivamento. "
                  "Responda APENAS com o texto transcrito, sem comentarios.")
        body = {"contents": [{"parts": [{"text": prompt}, {"inline_data": {"mime_type": "application/pdf", "data": pdf_b64}}]}]}
        url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-flash-latest:generateContent?key=" + GEMINI_KEY
        req = urllib.request.Request(url, data=json.dumps(body).encode(), headers={"Content-Type": "application/json"})
        resp = urllib.request.urlopen(req, timeout=40)
        data = json.loads(resp.read().decode())
        txt = data["candidates"][0]["content"]["parts"][0]["text"].strip()
        return txt
    except Exception as e:
        print("Gemini texto documento falhou:", e)
        return None

def _tesseract_texto_documento(caminho_pdf):
    import subprocess, os, glob, tempfile
    try:
        d = tempfile.mkdtemp()
        subprocess.run(["pdftoppm", "-r", "300", "-png", caminho_pdf, os.path.join(d, "pg")], check=True, timeout=90)
        texto = ""
        for img in sorted(glob.glob(os.path.join(d, "*.png"))):
            out = subprocess.run(["tesseract", img, "stdout", "-l", "por"], capture_output=True, text=True, timeout=60)
            if not out.stdout.strip():
                out = subprocess.run(["tesseract", img, "stdout"], capture_output=True, text=True, timeout=60)
            texto += out.stdout + "\n"
        return texto
    except Exception as e:
        print("Tesseract texto documento falhou:", e)
        return None

def _texto_parece_valido(texto: str) -> bool:
    """Detecta texto corrompido (fonte de PDF sem mapeamento Unicode correto).
    Um texto extraido corretamente deve ter alta proporcao de caracteres validos
    e conter palavras comuns em portugues."""
    import re
    if not texto or len(texto.strip()) < 50:
        return False
    validos = re.findall(r"[a-z0-9\u00e1\u00e0\u00e2\u00e3\u00e9\u00ea\u00ed\u00f3\u00f4\u00f5\u00fa\u00fc\u00e7A-Z\s.,;:()\-/\u00ba\u00aa%]", texto)
    proporcao_valida = len(validos) / max(len(texto), 1)
    if proporcao_valida < 0.85:
        return False
    t = texto.lower()
    palavras_comuns = [" de ", " da ", " do ", " que ", " para ", " com ", " em ", " uma ", " os ", " as ", " e "]
    if not any(p in t for p in palavras_comuns):
        return False
    return True

def _texto_printable_valido(texto, minimo=100):
    """Criterio simples pra camada 1 (pdfplumber): conta caracteres printaveis
    ASCII/Latin (letras, digitos, pontuacao, acentos comuns em portugues).
    Mais permissivo que _texto_parece_valido (que exige palavras inteiras em
    portugues) - usado especificamente na estrategia de leitura em camadas,
    onde a camada 1 so precisa confirmar que NAO e lixo binario/font corrompida."""
    if not texto:
        return False
    printaveis = sum(
        1 for c in texto
        if c.isprintable() and (c.isascii() or c in "áàâãéêíóôõúüçÁÀÂÃÉÊÍÓÔÕÚÜÇñÑ")
    )
    return printaveis > minimo


def _camada1_pdfplumber(caminho_pdf):
    """Camada 1 da estrategia de leitura de PDF (prioridade maxima - nunca
    bloquear insercao de processo por falha de leitura): extracao direta de
    texto via pdfplumber (biblioteca principal pedida) e, se vier curta,
    tambem via PyMuPDF/fitz como segunda tentativa da mesma camada (outra
    biblioteca de extracao direta, ja usada em producao e complementar -
    cada uma pode ter sucesso onde a outra falha, dependendo de como o PDF
    foi gerado). Retorna o texto se tiver mais de 100 caracteres legiveis,
    senao None (cai pra camada 2 - OCR)."""
    texto = ""
    try:
        import pdfplumber
        with pdfplumber.open(caminho_pdf) as pdf:
            for pagina in pdf.pages:
                texto += (pagina.extract_text() or "") + "\n"
    except Exception as e:
        print("   [PDF-camada1] pdfplumber falhou:", str(e)[:150])
        texto = ""

    # Alem da contagem de caracteres "legiveis" (_texto_printable_valido),
    # exige tambem _texto_parece_valido: fonte de PDF sem mapeamento Unicode
    # correto (comum em prints de pagina web) pode gerar uma sequencia de
    # tokens "(cid:123)" que passa facilmente dos 100 caracteres printaveis
    # (parenteses, letras, digitos - tudo ASCII) mas nao e texto de verdade -
    # sem essa checagem extra, esse lixo era aceito como camada 1 valida e a
    # camada 2 (OCR) nunca rodava, mesmo sendo o unico jeito de ler o PDF.
    if _texto_printable_valido(texto, minimo=100) and _texto_parece_valido(texto):
        print("   [PDF-camada1] pdfplumber OK,", len(texto.strip()), "caracteres")
        return texto

    try:
        import fitz
        doc = fitz.open(caminho_pdf)
        texto_fitz = ""
        for page in doc:
            texto_fitz += page.get_text()
        doc.close()
        if _texto_printable_valido(texto_fitz, minimo=100) and _texto_parece_valido(texto_fitz):
            print("   [PDF-camada1] fitz (2a tentativa) OK,", len(texto_fitz.strip()), "caracteres")
            return texto_fitz
    except Exception as e:
        print("   [PDF-camada1] fitz (2a tentativa) falhou:", str(e)[:150])

    print("   [PDF-camada1] extracao direta insuficiente - tentando camada 2 (OCR)")
    return None


def _camada2_ocr_pytesseract(caminho_pdf):
    """Camada 2 da estrategia de leitura de PDF: OCR. So chamada quando a
    camada 1 (extracao direta) falhou ou veio curta demais. Tenta primeiro
    Gemini vision (ja usado em producao, melhor em documentos escaneados
    reais com carimbo/selo) e, se indisponivel ou falhar, pdf2image +
    pytesseract com lang='por+eng' (biblioteca pedida). Retorna o texto
    resultante, ou None se todas as tentativas falharem."""
    texto_gemini = _gemini_texto_documento(caminho_pdf)
    if texto_gemini and texto_gemini.strip():
        print("   [PDF-camada2] OCR via Gemini vision OK,", len(texto_gemini.strip()), "caracteres")
        return texto_gemini.strip()

    try:
        from pdf2image import convert_from_path
        import pytesseract
        paginas = convert_from_path(caminho_pdf, dpi=300)
        texto = ""
        for img in paginas:
            texto += pytesseract.image_to_string(img, lang="por+eng") + "\n"
        texto = texto.strip()
        if texto:
            print("   [PDF-camada2] OCR pytesseract (por+eng) OK,", len(texto), "caracteres")
            return texto
    except Exception as e:
        print("   [PDF-camada2] OCR pytesseract falhou:", str(e)[:150])

    texto_tess_cli = _tesseract_texto_documento(caminho_pdf)
    if texto_tess_cli and texto_tess_cli.strip():
        print("   [PDF-camada2] OCR via tesseract CLI (fallback extra) OK,", len(texto_tess_cli.strip()), "caracteres")
        return texto_tess_cli.strip()

    print("   [PDF-camada2] nenhuma tentativa de OCR retornou texto")
    return None


def extrair_texto_pdf_em_camadas(caminho_pdf):
    """Estrategia de leitura de PDF em 3 camadas, em ordem de tentativa -
    PRIORIDADE MAXIMA: o sistema NUNCA pode falhar em inserir um processo por
    causa de um PDF dificil de ler.

    Camada 1 - pdfplumber (texto direto). Aceito se >100 caracteres legiveis.
    Camada 2 - OCR via pdf2image + pytesseract (lang='por+eng'), se a camada 1
      falhar ou vier curta/vazia demais.
    Camada 3 - fallback total: se o OCR tambem falhar ou vier com menos de 50
      caracteres, NAO bloqueia - devolve o que tiver (mesmo vazio/parcial) e
      sinaliza leitura_parcial=True pro chamador marcar o processo pra
      revisao manual do operador, sem nunca impedir a insercao.

    Retorna (texto: str, leitura_parcial: bool).
    """
    texto = _camada1_pdfplumber(caminho_pdf)
    if texto:
        return texto, False

    texto_ocr = _camada2_ocr_pytesseract(caminho_pdf)
    if texto_ocr and len(texto_ocr) >= 50:
        return texto_ocr, False

    # Camada 3: fallback total. Usa o melhor texto disponivel (mesmo curto),
    # nunca levanta excecao, nunca bloqueia - so sinaliza leitura parcial.
    melhor_texto = texto_ocr or ""
    print("   [PDF-camada3] leitura parcial/sem sucesso (", len(melhor_texto),
          "caracteres) - processo sera criado mesmo assim, marcado pra revisao manual")
    return melhor_texto, True


def _extrair_texto_bytes(conteudo: bytes, nome: str) -> str:
    import tempfile
    nm = (nome or "").lower()
    texto = ""
    try:
        if nm.endswith((".jpg", ".jpeg", ".png")):
            import img2pdf
            conteudo = img2pdf.convert(conteudo)
            nm = "convertida.pdf"
        if nm.endswith(".pdf"):
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as f:
                f.write(conteudo); tmp = f.name
            texto, _leitura_parcial = extrair_texto_pdf_em_camadas(tmp)
            print("DEBUG_EXTRACAO nome=", repr(nome), "| texto_len=", len(texto.strip()),
                  "| leitura_parcial=", _leitura_parcial, "| trecho=", repr(texto.strip()[:150]))
            os.unlink(tmp)
        elif nm.endswith(".docx"):
            import docx as docx_lib
            with tempfile.NamedTemporaryFile(delete=False, suffix=".docx") as f:
                f.write(conteudo); tmp = f.name
            doc = docx_lib.Document(tmp)
            texto = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
            os.unlink(tmp)
        else:
            texto = conteudo.decode("utf-8", errors="ignore")
    except Exception:
        texto = ""
    return texto

def _ja_registrada(texto_l: str) -> bool:
    """Detecta se o texto e' de uma ata JA REGISTRADA (comprovante de arquivamento),
    e portanto nunca deve ser tratada como documento principal (novo ato)."""
    tem_certifico = ("certifico o registro" in texto_l) or ("certifico o arquivamento" in texto_l)
    tem_sob_num = "sob o n" in texto_l
    return tem_certifico and tem_sob_num

def _classificar(nome: str, texto: str):
    import os as _os, unicodedata
    def _sa(s):
        s = (s or "").lower()
        return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    nome_l = _sa(nome)
    texto_l = _sa(texto[:4000])
    ext = _os.path.splitext((nome or "").lower())[1]
    score = 0
    tipo = None
    # PASSO 1 - nome do arquivo (prioridade)
    for t, marcs in TIPOS_PRINCIPAIS.items():
        for m in marcs:
            if _sa(m) in nome_l:
                score += 20
                if tipo is None:
                    tipo = t
    for m in MARCADORES_ANEXO:
        if _sa(m) in nome_l:
            score -= 12
    # PASSO 2 - conteudo (confirma/desempata)
    for t, marcs in TIPOS_PRINCIPAIS.items():
        for m in marcs:
            if _sa(m) in texto_l:
                score += 10
                if tipo is None:
                    tipo = t
    for m in MARCADORES_ANEXO:
        if _sa(m) in texto_l:
            score -= 4
    # PASSO 3 - imagem inclina a anexo (nao proibe)
    if ext in EXT_IMAGEM:
        score -= 6
    _jr = _ja_registrada(texto_l)
    print("DEBUG_JAREG nome=", repr(nome), "| ja_registrada=", _jr, "| trecho=", repr(texto_l[:300]))
    if _jr:
        tipo = None
        score -= 200
    return tipo, score

# ===== APRENDIZADO POR REGRAS ACUMULADAS =====
def _norm_ap(s):
    import unicodedata
    s = (s or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return " ".join(s.split())

def consultar_regras(nome, texto, db):
    """Retorna (classificacao, tipo_correto, peso) da melhor regra que casa, ou None."""
    base = _norm_ap((nome or "") + " " + (texto[:2000] or ""))
    if not base.strip():
        return None
    regras = db.query(RegraAprendizado).all()
    melhor = None
    for r in regras:
        padrao = _norm_ap(r.padrao)
        if padrao and padrao in base:
            if melhor is None or (r.peso or 1) > (melhor.peso or 1):
                melhor = r
    if melhor:
        return {"classificacao": melhor.classificacao, "tipo_correto": melhor.tipo_correto, "peso": melhor.peso}
    return None

@app.post("/aprendizado/registrar")
async def aprendizado_registrar(dados: str = Form(...), x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario or not usuario.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administrador")
    info = json.loads(dados)
    padrao = (info.get("padrao") or "").strip()
    classificacao = (info.get("classificacao") or "").strip()  # "principal" ou "anexo"
    tipo_correto = (info.get("tipo_correto") or "").strip()
    origem = (info.get("origem") or "nome").strip()
    if not padrao:
        raise HTTPException(status_code=400, detail="padrao obrigatorio")
    # se ja existe regra com mesmo padrao normalizado + classificacao, reforca o peso
    alvo = _norm_ap(padrao)
    existente = None
    for r in db.query(RegraAprendizado).all():
        if _norm_ap(r.padrao) == alvo and (r.classificacao or "") == classificacao and (r.tipo_correto or "") == tipo_correto:
            existente = r; break
    if existente:
        existente.peso = (existente.peso or 1) + 1
        db.commit()
        return {"mensagem": "Regra reforcada", "id": existente.id, "peso": existente.peso}
    nova = RegraAprendizado(id=str(uuid.uuid4()), padrao=padrao, origem=origem, classificacao=classificacao, tipo_correto=tipo_correto, peso=1, criado_por=usuario.login)
    db.add(nova)
    db.commit()
    return {"mensagem": "Regra criada", "id": nova.id}

@app.get("/aprendizado/regras")
async def aprendizado_listar(x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario or not usuario.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administrador")
    rs = db.query(RegraAprendizado).order_by(RegraAprendizado.peso.desc()).all()
    return [{"id": r.id, "padrao": r.padrao, "origem": r.origem, "classificacao": r.classificacao, "tipo_correto": r.tipo_correto, "peso": r.peso, "criado_por": r.criado_por} for r in rs]

@app.delete("/aprendizado/regras/{regra_id}")
async def aprendizado_apagar(regra_id: str, x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario or not usuario.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administrador")
    r = db.query(RegraAprendizado).filter(RegraAprendizado.id == regra_id).first()
    if r:
        db.delete(r); db.commit()
    return {"mensagem": "Regra removida"}

@app.post("/processos/analisar-pasta")
async def analisar_pasta(arquivos: list[UploadFile] = File(...), x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    itens = []
    for idx, arq in enumerate(arquivos):
        conteudo = await arq.read()
        texto = await asyncio.to_thread(_extrair_texto_bytes, conteudo, arq.filename or "")
        tipo, score = _classificar(arq.filename or "", texto)
        regra = consultar_regras(arq.filename or "", texto, db)
        regra_aplicada = None
        if regra:
            regra_aplicada = regra.get("classificacao")
            if regra.get("classificacao") == "principal":
                score += 100 + (regra.get("peso") or 1)
                if regra.get("tipo_correto"):
                    tipo = regra.get("tipo_correto")
            elif regra.get("classificacao") == "anexo":
                score -= 100 + (regra.get("peso") or 1)
        itens.append({"indice": idx, "nome": arq.filename or ("arquivo_" + str(idx)), "texto": texto, "tipo": tipo, "score": score, "regra_aplicada": regra_aplicada, "conteudo": conteudo})
    if not itens:
        raise HTTPException(status_code=400, detail="Nenhum arquivo recebido.")
    ordenados = sorted(itens, key=lambda x: x["score"], reverse=True)
    melhor = ordenados[0]
    maior = melhor["score"]
    bateram_principal = [i for i in itens if i["tipo"] is not None and i["score"] > 0]
    empatados_topo = [i for i in ordenados if i["score"] == maior]
    pendente = (len(bateram_principal) != 1) or (maior <= 0) or (len(empatados_topo) > 1)
    dados = analisar_ata_ia(melhor["texto"]) if melhor["texto"].strip() else {}
    dados["texto_extraido"] = melhor["texto"]
    if melhor["tipo"]:
        dados["tipo_ato"] = dados.get("tipo_ato") or melhor["tipo"]
    numero_prot = await asyncio.to_thread(_tentar_extrair_protocolo, melhor["conteudo"], melhor["nome"])
    if numero_prot:
        dados["numero_protocolo"] = numero_prot
    anexos = [{"indice": i["indice"], "nome": i["nome"]} for i in itens if i["indice"] != melhor["indice"]]
    return {
        "principal": {"indice": melhor["indice"], "nome": melhor["nome"], "tipo_sugerido": melhor["tipo"], "dados": dados, "score": melhor["score"]},
        "anexos": anexos,
        "confirmacao_pendente": pendente,
        "tipos_disponiveis": list(TIPOS_PRINCIPAIS.keys()),
        "candidatos": [{"indice": i["indice"], "nome": i["nome"], "tipo": i["tipo"], "score": i["score"]} for i in ordenados],
    }

def _filtrar_origem_destino(principais_out):
    """Se dois principais do mesmo lote forem a mesma empresa/ato/data mas UFs diferentes,
    e um apontar (uf_destino_transferencia) para a UF do outro, mantem so o de ORIGEM
    (o destino sera criado automaticamente depois, via _criar_processo_transferencia)."""
    def chave(item):
        d = item.get("dados") or {}
        return (_norm(d.get("empresa")), _norm(d.get("identificador_ato")), _norm(d.get("data_ata")))
    descartar = set()
    for i, a in enumerate(principais_out):
        for j, b in enumerate(principais_out):
            if i == j or i in descartar or j in descartar:
                continue
            if chave(a) != chave(b) or not chave(a)[0]:
                continue
            da = a.get("dados") or {}
            db_ = b.get("dados") or {}
            uf_a = (da.get("uf") or "").upper().strip()
            uf_b = (db_.get("uf") or "").upper().strip()
            dest_a = (da.get("uf_destino_transferencia") or "").upper().strip()
            dest_b = (db_.get("uf_destino_transferencia") or "").upper().strip()
            if dest_a and dest_a == uf_b:
                descartar.add(j)  # b e' o destino de a -> descarta b
            elif dest_b and dest_b == uf_a:
                descartar.add(i)  # a e' o destino de b -> descarta a
    return [p for k, p in enumerate(principais_out) if k not in descartar]


def _classificar_lote_ia(itens):
    """Classifica TODOS os documentos de um lote em uma unica chamada de IA,
    substituindo o antigo sistema de palavras-chave (fragil, quebra com qualquer
    variacao de titulo de documento nao prevista). Retorna dict {indice: {"principal": bool, "tipo_ato": str|None}}."""
    import json as _json, urllib.request
    if not GEMINI_KEY or not itens:
        return {}
    partes_doc = []
    for i in itens:
        trecho = (i["texto"] or "")[:3000]
        partes_doc.append(f"--- DOCUMENTO indice={i['indice']} nome=\"{i['nome']}\" ---\n{trecho}\n")
    prompt = (
        "Voce esta analisando um lote de documentos enviados para um sistema de gestao "
        "societaria brasileiro (Juntas Comerciais). Para CADA documento abaixo, determine:\n"
        "1) Se e um ATO PRINCIPAL (um documento que representa um ato societario formal que "
        "vira um processo proprio - ex: ata de assembleia/reuniao, alteracao contratual, "
        "contrato social, estatuto, distrato, protocolo de incorporacao, ata de resolucao de "
        "socio(a), qualquer documento assinado que registra uma deliberacao societaria) OU um "
        "ANEXO (documento de apoio - ex: identidade/CNH/RG de uma pessoa, procuracao, certidao, "
        "comprovante de pagamento, balanco, ficha cadastral, protocolo de junta comercial de "
        "OUTRO processo ja existente).\n"
        "2) Se for ANEXO, qual empresa/CNPJ ele parece se referir (para associa-lo ao ato principal certo).\n"
        "3) Se for PRINCIPAL, qual empresa/CNPJ esse documento se refere.\n\n"
        "Responda APENAS com um JSON no formato exato:\n"
        '{"classificacoes": [{"indice": 0, "principal": true, "empresa_ou_cnpj": "texto"}, '
        '{"indice": 1, "principal": false, "empresa_ou_cnpj": "texto"}]}\n\n'
        "Documentos:\n\n" + "\n".join(partes_doc)
    )
    body = {"contents": [{"parts": [{"text": prompt}]}]}
    url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-flash-latest:generateContent?key=" + GEMINI_KEY
    try:
        req = urllib.request.Request(url, data=_json.dumps(body).encode(), headers={"Content-Type": "application/json"})
        resp = urllib.request.urlopen(req, timeout=60)
        data = _json.loads(resp.read().decode())
        txt = data["candidates"][0]["content"]["parts"][0]["text"].strip()
        txt = txt.replace("```json", "").replace("```", "").strip()
        resultado = _json.loads(txt)
        saida = {}
        for c in resultado.get("classificacoes", []):
            saida[c["indice"]] = {"principal": bool(c.get("principal")), "empresa_ou_cnpj": c.get("empresa_ou_cnpj") or ""}
        return saida
    except Exception as e:
        print("Erro na classificacao por IA do lote:", str(e)[:200])
        return {}

def _tentar_extrair_protocolo(conteudo: bytes, nome: str):
    """Tenta extrair o numero de protocolo de um arquivo (PDF ou imagem), usando o
    mesmo pipeline robusto (codigo de barras + texto + Gemini + Tesseract) ja usado
    no upload de protocolo para processo existente. Nunca lanca excecao para fora."""
    import tempfile
    nm = (nome or "").lower()
    try:
        if nm.endswith((".jpg", ".jpeg", ".png")):
            import img2pdf
            conteudo = img2pdf.convert(conteudo)
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as f:
            f.write(conteudo)
            tmp = f.name
        numero = extrair_protocolo_ocr(tmp)
        os.unlink(tmp)
        return numero
    except Exception as e:
        print("Erro ao tentar extrair protocolo no upload de pasta:", str(e)[:150])
        return None
def _sa_texto_local(s):
    import unicodedata
    s = (s or "").lower()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

@app.post("/processos/analisar-pasta-multi")
async def analisar_pasta_multi(arquivos: list[UploadFile] = File(...), x_token: str = Header(None), db: Session = Depends(get_db)):
    """Detecta TODOS os documentos principais numa pasta/subpasta (nao so o melhor).
    Se houver mais de um principal, cada um vira um processo, e os demais arquivos
    (anexos) sao compartilhados/replicados entre todos os processos gerados."""
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    itens = []
    for idx, arq in enumerate(arquivos):
        conteudo = await arq.read()
        texto = await asyncio.to_thread(_extrair_texto_bytes, conteudo, arq.filename or "")
        ja_reg_flag = _ja_registrada(_sa_texto_local(texto[:4000]))
        itens.append({"indice": idx, "nome": arq.filename or ("arquivo_" + str(idx)), "texto": texto, "ja_reg": ja_reg_flag, "conteudo": conteudo})
    if not itens:
        raise HTTPException(status_code=400, detail="Nenhum arquivo recebido.")

    # CLASSIFICACAO POR IA (uma chamada para o lote inteiro) - substitui o antigo
    # sistema de palavras-chave, que quebrava com qualquer variacao de titulo nao
    # prevista. Documentos ja registrados nunca sao candidatos.
    candidatos_ia = [i for i in itens if not i.get("ja_reg")]
    classificacao_ia = await asyncio.to_thread(_classificar_lote_ia, candidatos_ia)

    principais_itens = []
    pendente = False
    for i in candidatos_ia:
        c = classificacao_ia.get(i["indice"])
        if c is None:
            principais_itens.append(i)
            pendente = True
        elif c.get("principal"):
            principais_itens.append(i)

    indices_principais = {i["indice"] for i in principais_itens}
    anexos = [{"indice": i["indice"], "nome": i["nome"]} for i in itens if i["indice"] not in indices_principais]

    principais_out = []
    for i in principais_itens:
        # Camada 3 da leitura de PDF (prioridade maxima): texto insuficiente
        # (<50 caracteres) nunca bloqueia - so marca leitura_parcial=True pra
        # criar_processo sinalizar revisao manual do operador.
        leitura_parcial = len((i["texto"] or "").strip()) < 50
        dados = analisar_ata_ia(i["texto"]) if i["texto"].strip() else dict(_CAMPOS_VAZIOS_ATA)
        dados["leitura_parcial"] = leitura_parcial
        dados["texto_extraido"] = i["texto"]
        numero_prot = await asyncio.to_thread(_tentar_extrair_protocolo, i["conteudo"], i["nome"])
        if numero_prot:
            dados["numero_protocolo"] = numero_prot
        principais_out.append({"indice": i["indice"], "nome": i["nome"], "tipo_sugerido": dados.get("tipo_ato"), "dados": dados, "score": 0})

    for _pp in principais_out:
        _dd = _pp.get("dados") or {}
        print("DEBUG_DEDUP antes: nome=", _pp.get("nome"), "| empresa=", _dd.get("empresa"), "| ato=", _dd.get("identificador_ato"), "| data=", _dd.get("data_ata"), "| uf=", _dd.get("uf"), "| uf_destino=", _dd.get("uf_destino_transferencia"))
    principais_out = _filtrar_origem_destino(principais_out)
    print("DEBUG_DEDUP depois: total=", len(principais_out))

    # Fallback: comprovante de protocolo (JUCESP/JUCERJA) costuma vir como PDF
    # separado da ata no mesmo lote (print de pagina web) e a classificacao por
    # IA o marca como anexo, nao principal - nesse caso o protocolo nunca era
    # procurado nele. So aplicado quando ha exatamente 1 processo principal no
    # lote (evita ambiguidade de a qual processo o protocolo pertence).
    if len(principais_out) == 1 and not (principais_out[0]["dados"].get("numero_protocolo") or "").strip():
        indices_principais_final = {p["indice"] for p in principais_out}
        for i in itens:
            if i["indice"] in indices_principais_final:
                continue
            numero_prot_anexo = await asyncio.to_thread(_tentar_extrair_protocolo, i["conteudo"], i["nome"])
            if numero_prot_anexo:
                principais_out[0]["dados"]["numero_protocolo"] = numero_prot_anexo
                print("Protocolo extraido de anexo do lote:", numero_prot_anexo, "arquivo:", i["nome"])
                break

    return {
        "principais": principais_out,
        "anexos": anexos,
        "multiplo": len(principais_out) > 1,
        "confirmacao_pendente": pendente,
        "tipos_disponiveis": list(TIPOS_PRINCIPAIS.keys()),
    }

@app.post("/processos/{processo_id}/confirmar-tipo")
async def confirmar_tipo(processo_id: str, dados: str = Form(...), request: Request = None, x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    requer_acesso_admin(usuario)
    p = db.query(Processo).filter(Processo.id == processo_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")
    info = json.loads(dados)
    novo_tipo = (info.get("tipo_ato") or "").strip()
    if novo_tipo:
        p.tipo_ato = novo_tipo
    p.confirmacao_pendente = False
    db.commit()
    _ip = obter_ip(request)
    registrar_auditoria(db, usuario, "confirmar_tipo", processo_id, "tipo=" + (novo_tipo or p.tipo_ato or ""), _ip)
    return {"mensagem": "Tipo confirmado", "id": processo_id, "tipo_ato": p.tipo_ato}

def _norm(s):
    import unicodedata
    s = (s or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return " ".join(s.split())

@app.post("/processos/analisar")
async def analisar_documento(arquivo: UploadFile = File(...), x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")

    conteudo = await arquivo.read()
    nome = arquivo.filename or ""
    texto = await asyncio.to_thread(_extrair_texto_bytes, conteudo, nome)

    dados = analisar_ata_ia(texto) if texto.strip() else dict(_CAMPOS_VAZIOS_ATA)
    dados["leitura_parcial"] = len((texto or "").strip()) < 50
    dados["texto_extraido"] = texto
    return dados

@app.post("/processos")
async def criar_processo(
    arquivo: UploadFile = File(None),
    dados: str = Form(...),
    x_token: str = Header(None),
    db: Session = Depends(get_db)
):
    info = json.loads(dados)
    processo_id = f"MN-{datetime.now().strftime('%Y%m%d%H%M%S')}-{str(uuid.uuid4())[:4].upper()}"

    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario_tok = validar_token(x_token, db)
    if not usuario_tok:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")

    obrigatorios = {
        "empresa": (info.get("empresa") or "").strip(),
        "tipo_ato": (info.get("tipo_ato") or "").strip(),
        "data_ata": (info.get("data_ata") or "").strip(),
    }
    faltando = [campo for campo, valor in obrigatorios.items() if not valor]
    # URGENTE - NUNCA bloquear a insercao do processo por falta de campo extraido.
    # Sempre insere o processo, marca para revisao manual, e avisa o administrador.

    cnpj_norm = normalizar_cnpj(info.get("cnpj") or "")
    if cnpj_norm and not validar_cnpj(cnpj_norm):
        faltando.append("cnpj (digito verificador invalido)")
    cnpj_final = formatar_cnpj(cnpj_norm) if cnpj_norm else ""

    grupo_id = None
    if _tem_acesso_admin(usuario_tok):
        codigo_grupo = info.get("codigo_grupo", "").strip()
        if codigo_grupo:
            grupo = db.query(Grupo).filter(Grupo.codigo == codigo_grupo).first()
            if not grupo:
                raise HTTPException(status_code=400, detail=f"Grupo com codigo '{codigo_grupo}' nao encontrado")
            grupo_id = grupo.id
    else:
        grupo_id = usuario_tok.grupo_id

    arquivo_ata = None
    if arquivo:
        ext = os.path.splitext(arquivo.filename)[1]
        nome_arquivo = f"{processo_id}_ata{ext}"
        caminho = os.path.join(UPLOADS_DIR, nome_arquivo)
        with open(caminho, "wb") as f:
            f.write(await arquivo.read())
        arquivo_ata = nome_arquivo

    p = Processo(
        id=processo_id,
        empresa=(info.get("empresa", "") or "").upper(),
        cnpj=cnpj_final,
        nire=info.get("nire", ""),
        uf=(info.get("uf") or "").upper().strip()[:2],
        tipo_sociedade=info.get("tipo_sociedade", ""),
        tipo_ato=info.get("tipo_ato", ""),
        identificador_ato=info.get("identificador_ato", ""),
        data_ata=info.get("data_ata", ""),
        hora_ata=info.get("hora_ata", ""),
        numero_protocolo=info.get("numero_protocolo", ""),
        email_cliente=info.get("email_cliente", ""),
        eventos=json.dumps(info.get("eventos", []), ensure_ascii=False),
        checklist=json.dumps(info.get("checklist", []), ensure_ascii=False),
        requer_cpl=info.get("requer_cpl", False),
        observacoes=info.get("observacoes", ""),
        status="aberto",
        arquivo_ata=arquivo_ata,
        grupo_id=grupo_id,
        uf_destino_transferencia=(info.get("uf_destino_transferencia") or "").upper().strip()[:2] or None,
        leitura_parcial=bool(info.get("leitura_parcial", False)),
        texto_documento_extraido=info.get("texto_extraido") or None,
    )
    db.add(p)
    db.flush()
    vincular_fluxo_do_dia(db, p, grupo_id)
    registrar_evento(db, p, "ata_enviada", "Ata enviada", usuario_tok)
    db.commit()
    try:
        corpo = "Processo Inserido no Atos:\n\n" + corpo_status_cliente(p, "Aberto", "")
        try:
            _recebido_txt = p.data_recebimento.strftime("%d/%m/%Y, %H:%M")
        except Exception:
            _recebido_txt = datetime.now().strftime("%d/%m/%Y, %H:%M")
        corpo_html = _email_status_html("aberto", "Aberto", "Seu processo foi recebido", _empresa_linha(p),
                                         nota_tipo="recebido", nota_texto=_recebido_txt,
                                         botao={"label": "Acessar o sistema", "href": BASE_URL_SISTEMA})
        for em in emails_do_grupo(db, grupo_id):
            enviar_email(em, "Processo inserido no Atos - " + (p.empresa or ""), corpo, corpo_html)
    except Exception as e:
        print("Erro ao notificar abertura:", e)
    if faltando:
        try:
            p.confirmacao_pendente = True
            db.commit()
            _assunto_incompleto = "[Atos] ATENCAO - Processo inserido com campos incompletos - " + (p.empresa or processo_id)
            _corpo_incompleto = "O processo " + processo_id + " (" + (p.empresa or "sem nome") + ") foi inserido no sistema, mas a extracao automatica nao conseguiu identificar: " + ", ".join(faltando) + ".\n\nRevise manualmente e complete os dados faltantes o quanto antes."
            for _e in emails_admin(db):
                enviar_email(_e, _assunto_incompleto, _corpo_incompleto)
        except Exception as e:
            print("Erro ao notificar campos incompletos:", e)
    if p.leitura_parcial:
        try:
            p.confirmacao_pendente = True
            db.commit()
            _assunto_leitura = "[Atos] ATENCAO - Leitura parcial de PDF - " + (p.empresa or processo_id)
            _corpo_leitura = "O processo " + processo_id + " (" + (p.empresa or "sem nome") + ") foi inserido no sistema, mas a leitura do PDF (texto direto + OCR) nao conseguiu extrair conteudo suficiente.\n\nO processo entrou normalmente no sistema, mas revise manualmente o documento e complete/corrija os dados o quanto antes."
            for _e in emails_admin(db):
                enviar_email(_e, _assunto_leitura, _corpo_leitura)
        except Exception as e:
            print("Erro ao notificar leitura parcial:", e)
    return {"id": processo_id, "mensagem": "Processo criado com sucesso"}

def _criar_processo_transferencia(db, p_origem):
    """Cria automaticamente o processo de destino apos a origem ser finalizada,
    quando a ata identificou transferencia de sede interestadual."""
    uf_destino = (p_origem.uf_destino_transferencia or "").strip().upper()
    if not uf_destino or p_origem.transferencia_criada:
        return None
    novo_id = f"MN-{datetime.now().strftime('%Y%m%d%H%M%S')}-{str(uuid.uuid4())[:4].upper()}"
    obs = f"Processo criado automaticamente apos transferencia de sede. Origem: {p_origem.id} ({p_origem.uf or '-'})."
    novo = Processo(
        id=novo_id,
        empresa=(p_origem.empresa or "").upper(),
        cnpj=p_origem.cnpj,
        nire=p_origem.nire,
        uf=uf_destino,
        tipo_sociedade=p_origem.tipo_sociedade,
        tipo_ato=p_origem.tipo_ato,
        identificador_ato=(p_origem.identificador_ato or "") + " - Transferencia de Sede (Destino)",
        data_ata=p_origem.data_ata,
        hora_ata=p_origem.hora_ata,
        email_cliente=p_origem.email_cliente,
        observacoes=obs,
        status="aberto",
        grupo_id=p_origem.grupo_id,
        processo_origem_id=p_origem.id,
    )
    db.add(novo)
    db.flush()
    vincular_fluxo_do_dia(db, novo, novo.grupo_id)
    registrar_evento(db, novo, "processo_criado_transferencia", "Processo de destino criado automaticamente (transferência de sede)")
    # anexa a ata de origem (ja registrada) como comprovante no processo novo
    if p_origem.arquivo_ata:
        try:
            origem_path = os.path.join(UPLOADS_DIR, p_origem.arquivo_ata)
            if os.path.exists(origem_path):
                ext = os.path.splitext(p_origem.arquivo_ata)[1]
                anexo_id = str(uuid.uuid4())
                nome_anexo = "anexo_" + anexo_id + ext
                destino_path = os.path.join(UPLOADS_DIR, nome_anexo)
                with open(origem_path, "rb") as fr, open(destino_path, "wb") as fw:
                    fw.write(fr.read())
                db.add(Anexo(
                    id=anexo_id, processo_id=novo_id, arquivo=nome_anexo,
                    nome_original="Ata registrada (origem " + (p_origem.uf or "") + ")",
                    descricao="Comprovante de registro na Junta de origem, anexado automaticamente.",
                    enviado_por="sistema",
                ))
        except Exception as e:
            print("Erro ao anexar ata de origem no processo de transferencia:", e)
    p_origem.transferencia_criada = True
    db.commit()
    try:
        notificar_telegram(f"ATOS - Transferencia de sede\nProcesso de destino criado: {novo_id}\nEmpresa: {p_origem.empresa}\nDestino: {uf_destino}\nAguardando protocolo.")
    except Exception:
        pass
    return novo_id

def recalcular_status(p):
    # Prioridade: registro(finalizado) > exigencia > deferido(automacao) > protocolo(tramitacao) > aberto
    if p.arquivo_registro:
        return "finalizado"
    if getattr(p, "exigencia_ativa", False):
        return "exigencia"
    if (p.status or "").lower() == "deferido":
        return "deferido"
    if p.numero_protocolo or p.arquivo_protocolo:
        return "tramitacao"
    return "aberto"


@app.patch("/processos/{processo_id}")
def atualizar_processo(processo_id: str, dados: dict, request: Request = None, x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    p = db.query(Processo).filter(Processo.id == processo_id).first()
    if p and not _tem_acesso_admin(usuario) and p.grupo_id != usuario.grupo_id:
        raise HTTPException(status_code=403, detail="Sem permissao para este processo")
    if not p:
        raise HTTPException(status_code=404, detail="Processo não encontrado")
    _ip = obter_ip(request)
    registrar_auditoria(db, usuario, "editar", processo_id, "campos=" + ",".join(list(dados.keys())), _ip)
    for campo, valor in dados.items():
        if hasattr(p, campo):
            if campo == "empresa" and valor:
                valor = valor.upper()
            if campo == "cnpj" and valor:
                valor = formatar_cnpj(normalizar_cnpj(valor))
            setattr(p, campo, valor)
    # Reinserir/atualizar protocolo cumpre a exigencia ativa
    protocolo_editado = "numero_protocolo" in dados or "arquivo_protocolo" in dados
    if protocolo_editado and getattr(p, "exigencia_ativa", False):
        p.exigencia_ativa = False
    status_antes_patch = (p.status or "").lower()
    p.status = recalcular_status(p)
    p.atualizado_em = datetime.now()
    if protocolo_editado:
        registrar_evento(db, p, "protocolo_inserido", "Protocolo inserido manualmente" + (f": {p.numero_protocolo}" if p.numero_protocolo else ""), usuario)
    db.commit()
    notificar_tramitacao_cliente(db, p, status_antes_patch)
    return {"mensagem": "Atualizado com sucesso"}

@app.post("/processos/{processo_id}/upload/{tipo}")
async def upload_arquivo(
    processo_id: str,
    tipo: str,
    arquivo: UploadFile = File(...),
    request: Request = None,
    x_token: str = Header(None),
    db: Session = Depends(get_db)
):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    p = db.query(Processo).filter(Processo.id == processo_id).first()
    if p and not _tem_acesso_admin(usuario) and p.grupo_id != usuario.grupo_id:
        raise HTTPException(status_code=403, detail="Sem permissao para este processo")
    if not p:
        raise HTTPException(status_code=404, detail="Processo não encontrado")

    ext = os.path.splitext(arquivo.filename or "")[1].lower()
    # validacao: so extensoes permitidas
    EXT_PERMITIDAS = {".pdf", ".png", ".jpg", ".jpeg"}
    if ext not in EXT_PERMITIDAS:
        raise HTTPException(status_code=400, detail="Tipo de arquivo nao permitido. Envie PDF ou imagem.")
    # validacao: tamanho maximo 20 MB
    conteudo = await arquivo.read()
    MAX_BYTES = 20 * 1024 * 1024
    if len(conteudo) > MAX_BYTES:
        raise HTTPException(status_code=400, detail="Arquivo muito grande. Limite de 20 MB.")
    if len(conteudo) == 0:
        raise HTTPException(status_code=400, detail="Arquivo vazio.")
    # validacao: se diz ser PDF, conferir a assinatura real do arquivo
    if ext == ".pdf" and not conteudo[:5].startswith(b"%PDF"):
        raise HTTPException(status_code=400, detail="Arquivo nao e um PDF valido.")
    nome_arquivo = f"{processo_id}_{tipo}{ext}"
    caminho = os.path.join(UPLOADS_DIR, nome_arquivo)
    with open(caminho, "wb") as f:
        f.write(conteudo)
    _ip = obter_ip(request)
    registrar_auditoria(db, usuario, "upload", processo_id, "tipo=" + str(tipo) + " arquivo=" + str(nome_arquivo), _ip)

    campo_map = {
        "protocolo": "arquivo_protocolo",
        "registro": "arquivo_registro",
        "nd": "arquivo_nd",
        "nf": "arquivo_nf"
    }
    if tipo in campo_map:
        status_antes_up = (p.status or "").lower()
        setattr(p, campo_map[tipo], nome_arquivo)
        if tipo == "protocolo":
            _num = await asyncio.to_thread(_tentar_extrair_protocolo, conteudo, nome_arquivo)
            if _num:
                p.numero_protocolo = _num
                print("OCR protocolo detectado:", _num)
        if tipo == "protocolo" and getattr(p, "exigencia_ativa", False):
            p.exigencia_ativa = False
        p.status = recalcular_status(p)
        p.atualizado_em = datetime.now()
        _evento_upload = {
            "protocolo": ("protocolo_inserido", "Protocolo inserido" + (f": {p.numero_protocolo}" if p.numero_protocolo else "")),
            "registro": ("registro_finalizado", "Ata registrada"),
            "nd": ("nd_inserida", "Nota de Débito inserida"),
            "nf": ("nf_inserida", "Nota Fiscal inserida"),
        }.get(tipo)
        if _evento_upload:
            registrar_evento(db, p, _evento_upload[0], _evento_upload[1], usuario)
        db.commit()
        try:
            novo_status = (p.status or "").lower()
            if tipo == "registro" and novo_status == "finalizado":
                # SUSPENSO 30/07/2026: e-mail ao cliente da automacao JUCESP
                # (SP) desativado por decisao explicita, ate revisao do fluxo
                # de documento. Upload/registro continua funcionando
                # normalmente (arquivo salvo, status atualizado) - so o aviso
                # ao cliente que fica em pausa pra SP.
                if (p.uf or "").upper() != "SP":
                    corpo, corpo_html = _email_finalizado(p)
                    for em in emails_do_grupo(db, p.grupo_id):
                        enviar_email_anexo(em, "Processo Finalizado - " + (p.empresa or ""), corpo, caminho, nome_arquivo, corpo_html=corpo_html)
                try:
                    _criar_processo_transferencia(db, p)
                except Exception as e:
                    print("Erro ao criar processo de transferencia:", e)
            if tipo == "protocolo":
                notificar_tramitacao_cliente(db, p, status_antes_up)
        except Exception as e:
            print("Erro ao notificar upload:", e)

    return {"mensagem": f"Arquivo {tipo} salvo", "arquivo": nome_arquivo, "numero_protocolo": (p.numero_protocolo or "")}

@app.post("/processos/{processo_id}/exigencia")
async def registrar_exigencia(
    processo_id: str,
    texto: str = Form(""),
    arquivo: UploadFile = File(None),
    x_token: str = Header(None),
    db: Session = Depends(get_db)
):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    p = db.query(Processo).filter(Processo.id == processo_id).first()
    if p and not _tem_acesso_admin(usuario) and p.grupo_id != usuario.grupo_id:
        raise HTTPException(status_code=403, detail="Sem permissao para este processo")
    if not p:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")
    p.texto_exigencia = texto
    if arquivo is not None:
        ext = os.path.splitext(arquivo.filename)[1]
        nome_arquivo = f"{processo_id}_exigencia{ext}"
        caminho = os.path.join(UPLOADS_DIR, nome_arquivo)
        with open(caminho, "wb") as f:
            f.write(await arquivo.read())
        p.arquivo_exigencia = nome_arquivo
    p.exigencia_ativa = True
    p.status = recalcular_status(p)
    p.atualizado_em = datetime.now()
    registrar_evento(db, p, "exigencia_registrada", "Exigência registrada" + (f": {texto}" if texto else ""), usuario)
    db.commit()
    notificar_exigencia_cliente(db, p, origem="manual")
    return {"mensagem": "Exigencia registrada", "status": p.status}


@app.post("/processos/{processo_id}/exigencia/cumprida")
def exigencia_cumprida(processo_id: str, x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    p = db.query(Processo).filter(Processo.id == processo_id).first()
    if p and not _tem_acesso_admin(usuario) and p.grupo_id != usuario.grupo_id:
        raise HTTPException(status_code=403, detail="Sem permissao para este processo")
    if not p:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")
    p.exigencia_ativa = False
    p.status = recalcular_status(p)
    p.atualizado_em = datetime.now()
    registrar_evento(db, p, "exigencia_cumprida", "Exigência marcada como cumprida", usuario)
    db.commit()
    return {"mensagem": "Exigencia marcada como cumprida", "status": p.status}

@app.delete("/processos/{processo_id}")
def excluir_processo(processo_id: str, request: Request = None, x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario or not usuario.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administrador pode excluir processos")
    p = db.query(Processo).filter(Processo.id == processo_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")
    _ip = obter_ip(request)
    registrar_auditoria(db, usuario, "excluir", processo_id, "empresa=" + str(p.empresa) + " cnpj=" + str(p.cnpj), _ip)
    db.delete(p)
    db.commit()
    return {"mensagem": "Processo excluido"}

@app.post("/processos/{processo_id}/exigencia/aguardando-cliente")
def exigencia_aguardando_cliente(processo_id: str, x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    requer_acesso_admin(usuario)
    p = db.query(Processo).filter(Processo.id == processo_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")
    p.aguardando_cliente = True
    p.atualizado_em = datetime.now()
    db.commit()
    return {"mensagem": "Marcado como aguardando cliente", "aguardando_cliente": True}


@app.get("/grupos")
def listar_grupos(x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    requer_acesso_admin(usuario)
    grupos = db.query(Grupo).order_by(Grupo.nome).all()
    return [{"id": g.id, "nome": g.nome, "codigo": g.codigo} for g in grupos]

@app.post("/grupos/criar")
def criar_grupo(dados: dict, background: BackgroundTasks, x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    requer_acesso_admin(usuario)

    nome = (dados.get("nome") or "").strip()
    emails = dados.get("emails") or []
    emails = [e.strip() for e in emails if e and e.strip()]

    if not nome:
        raise HTTPException(status_code=400, detail="Informe o nome do grupo")
    if not emails:
        raise HTTPException(status_code=400, detail="Informe ao menos um email")

    existente = db.query(Grupo).filter(Grupo.nome == nome).first()
    if existente:
        raise HTTPException(status_code=400, detail=f"Ja existe um grupo chamado '{nome}'")

    base = "".join(ch for ch in nome.upper() if ch.isalnum())[:8] or "GRUPO"
    codigo = f"{base}-{uuid.uuid4().hex[:4].upper()}"
    grupo = Grupo(id=str(uuid.uuid4()), nome=nome, codigo=codigo)
    db.add(grupo)
    db.commit()

    link = f"{BASE_URL_SISTEMA}/cliente?grupo={codigo}"
    for email in emails:
        ja = db.query(EmailGrupo).filter(EmailGrupo.email == email, EmailGrupo.grupo_id == grupo.id).first()
        if not ja:
            db.add(EmailGrupo(id=str(uuid.uuid4()), email=email, grupo_id=grupo.id))
            db.commit()
    background.add_task(_disparar_convites, nome, link, emails)
    enviados = emails
    falharam = []

    return {
        "mensagem": "Grupo criado com sucesso",
        "grupo": nome,
        "codigo": codigo,
        "emails_enviados": enviados,
        "emails_falharam": falharam,
        "link": link
    }


TOKEN_CONVITE_VALIDADE_HORAS = 48


def gerar_convite(db, usuario, horas_validade=TOKEN_CONVITE_VALIDADE_HORAS):
    """Gera token de convite de uso unico pro usuario definir a propria senha
    no primeiro acesso (ou redefinir apos convite reenviado). Sobrescreve
    qualquer convite anterior ainda pendente."""
    token = secrets.token_urlsafe(32)
    usuario.token_convite = token
    usuario.convite_expira_em = datetime.now() + timedelta(hours=horas_validade)
    db.commit()
    return token


def enviar_convite_email(usuario, token):
    link = f"{BASE_URL_SISTEMA}/criar-senha?token={token}"
    corpo = (
        "Ola, " + (usuario.nome or usuario.login) + "!\n\n"
        "Voce foi convidado a acessar o Atos - Gestao Societaria.\n\n"
        "Clique no link abaixo para definir sua senha de acesso (valido por "
        + str(TOKEN_CONVITE_VALIDADE_HORAS) + " horas):\n" + link + "\n\n"
        "Se voce nao esperava este e-mail, ignore-o."
    )
    return enviar_email(usuario.email, "Convite de acesso - Atos", corpo)


@app.get("/convite/{token}")
def validar_convite(token: str, db: Session = Depends(get_db)):
    """Endpoint publico (sem token de sessao) pra tela /criar-senha conferir se
    o link ainda e valido antes de mostrar o formulario, e cumprimentar pelo nome."""
    usuario = db.query(Usuario).filter(Usuario.token_convite == token).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Convite invalido ou ja utilizado")
    if usuario.convite_expira_em and usuario.convite_expira_em < datetime.now():
        raise HTTPException(status_code=410, detail="Convite expirado. Peca um novo convite ao administrador.")
    return {"nome": usuario.nome or usuario.login, "email": usuario.email}


@app.post("/convite/definir-senha")
def definir_senha_convite(dados: dict, db: Session = Depends(get_db)):
    """Endpoint publico: define a senha a partir de um token de convite valido
    e invalida o token (uso unico) - mesma tela usada tanto pro primeiro acesso
    quanto pra um convite reenviado."""
    token = (dados.get("token") or "").strip()
    senha = dados.get("senha") or ""
    if not token or not senha:
        raise HTTPException(status_code=400, detail="token e senha sao obrigatorios")
    if len(senha) < 8 or not any(c.isalpha() for c in senha) or not any(c.isdigit() for c in senha):
        raise HTTPException(status_code=400, detail="A senha deve ter ao menos 8 caracteres, com letras e numeros")

    usuario = db.query(Usuario).filter(Usuario.token_convite == token).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Convite invalido ou ja utilizado")
    if usuario.convite_expira_em and usuario.convite_expira_em < datetime.now():
        raise HTTPException(status_code=410, detail="Convite expirado. Peca um novo convite ao administrador.")

    usuario.senha_hash = bcrypt.hashpw(senha.encode()[:72], bcrypt.gensalt()).decode()
    usuario.token_convite = None
    usuario.convite_expira_em = None
    db.commit()
    return {"mensagem": "Senha definida com sucesso. Voce ja pode fazer login."}


@app.post("/usuarios/operador")
def criar_usuario_operador(dados: dict, request: Request = None, x_token: str = Header(None), db: Session = Depends(get_db)):
    """Cria um novo usuario com papel 'operador' (acesso a tela administrativa,
    sem configuracoes/gerenciamento de usuarios/identidade visual). So o admin
    completo pode criar. Nao recebe senha - o usuario define a propria senha
    via convite por e-mail (fluxo /criar-senha), mesmo padrao usado pra
    convidar/reconvidar qualquer usuario administrativo."""
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    requer_admin_completo(usuario)

    nome = (dados.get("nome") or "").strip()
    email = (dados.get("email") or "").strip().lower()
    login = (dados.get("login") or email).strip()

    if not nome or not email:
        raise HTTPException(status_code=400, detail="nome e email sao obrigatorios")

    existente = db.query(Usuario).filter(Usuario.login == login).first()
    if existente:
        raise HTTPException(status_code=400, detail="Ja existe um usuario com esse login")

    grupo = db.query(Grupo).filter(Grupo.codigo == "ADMIN").first()
    if not grupo:
        grupo = Grupo(id=str(uuid.uuid4()), nome="Administracao", codigo="ADMIN")
        db.add(grupo)
        db.commit()

    # senha_hash placeholder (aleatorio, nunca comunicado) - login so funciona
    # depois que o convite definir uma senha de verdade.
    senha_placeholder = bcrypt.hashpw(secrets.token_urlsafe(24).encode()[:72], bcrypt.gensalt()).decode()
    novo = Usuario(
        id=str(uuid.uuid4()),
        login=login,
        senha_hash=senha_placeholder,
        email=email,
        nome=nome,
        grupo_id=grupo.id,
        is_admin=False,
        papel="operador",
    )
    db.add(novo)
    db.commit()

    token = gerar_convite(db, novo)
    enviar_convite_email(novo, token)

    _ip = obter_ip(request)
    registrar_auditoria(db, usuario, "criar_operador", None, "novo_login=" + login + " nome=" + nome, _ip)
    return {"mensagem": "Usuario operador criado, convite enviado por e-mail", "login": login, "nome": nome, "email": email}


@app.get("/relatorio")
def gerar_relatorio(status: str = "todos", x_token: str = Header(None), db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from fastapi.responses import StreamingResponse
    import io
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido")
    query = db.query(Processo).filter(Processo.grupo_id == usuario.grupo_id)
    if status and status != "todos":
        query = query.filter(Processo.status == status)
    from sqlalchemy import case
    processos = query.order_by(
        case((Processo.status == "finalizado", 1), else_=0),
        case((Processo.status == "finalizado", Processo.atualizado_em), else_=Processo.criado_em).desc()
    ).all()
    rotulos = {"recebido": "Aberto", "tramitacao": "Tramitacao", "exigencia": "Exigencia", "aprovado": "Deferido"}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Processos"
    cabecalho = ["Empresa", "CNPJ", "UF", "Ato", "Protocolo", "Status"]
    ws.append(cabecalho)
    for c in range(1, len(cabecalho) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4D52")
    for p in processos:
        ws.append([
            p.empresa or "",
            p.cnpj or "",
            p.uf or "",
            p.identificador_ato or p.tipo_ato or "",
            (p.numero_protocolo or ""),
            rotulos.get(p.status, p.status or ""),
        ])
    larguras = [42, 22, 6, 50, 16, 14]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome = f"relatorio_{status}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={nome}"})


@app.get("/metricas")
def metricas(x_token: str = Header(None), db: Session = Depends(get_db)):
    if not x_token:
        raise HTTPException(status_code=401, detail="Token necessario")
    usuario = validar_token(x_token, db)
    if not usuario:
        raise HTTPException(status_code=401, detail="Token invalido ou sessao expirada")
    base = db.query(Processo)
    if not _tem_acesso_admin(usuario):
        base = base.filter(Processo.grupo_id == usuario.grupo_id)
    total = base.count()
    tramitacao = base.filter(Processo.status == "tramitacao").count()
    exigencia = base.filter(Processo.status == "exigencia").count()
    aprovado = base.filter(Processo.status == "aprovado").count()
    deferido = base.filter(Processo.status == "deferido").count()
    finalizado = base.filter(Processo.status == "finalizado").count()
    cobranca_pendente = base.filter(
        Processo.status.in_(["aprovado", "finalizado"]),
        Processo.nf_enviada == False
    ).count()
    return {
        "total": total,
        "tramitacao": tramitacao,
        "exigencia": exigencia,
        "aprovado": aprovado,
        "deferido": deferido,
        "finalizado": finalizado,
        "cobranca_pendente": cobranca_pendente
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)